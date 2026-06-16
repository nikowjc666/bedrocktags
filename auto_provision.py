"""
Bedrock Inference Profile 一键自动配置工具
==========================================
支持: 创建配置+打标签+测试 | 删除配置 | 导出 Excel | Watch 守护模式

使用方式:
  1. Web UI:  python auto_provision.py  → 浏览器打开 http://localhost:5050
  2. CLI 创建: python auto_provision.py --cli --link-id 621724235498 \
              --models "4.5,4.6,4.7,4.8,fable 5" --regions "美西,美东" \
              --tag-key "migrationId" --tag-value "migWEII3463IB" \
              --ak YOUR_AK --sk YOUR_SK
  3. CLI 删除: python auto_provision.py --cli --delete --link-id 621724235498 \
              --regions "美西,美东" --ak YOUR_AK --sk YOUR_SK
  4. Watch:   python auto_provision.py --watch
"""

import json
import time
import re
import sys
import argparse
import os
import boto3
import botocore
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ═══════════════════════════════════════════════════════════════════
# 配置常量
# ═══════════════════════════════════════════════════════════════════

# 中文简写 → AWS Region 映射
REGION_ALIAS = {
    "美东": ["us-east-1"],
    "美东1": ["us-east-1"],
    "美东2": ["us-east-2"],
    "美西": ["us-west-2"],
    "美西1": ["us-west-1"],
    "美西2": ["us-west-2"],
    "欧洲": ["eu-west-1"],
    "欧西": ["eu-west-1"],
    "欧中": ["eu-central-1"],
    "法兰克福": ["eu-central-1"],
    "爱尔兰": ["eu-west-1"],
    "伦敦": ["eu-west-2"],
    "巴黎": ["eu-west-3"],
    "东京": ["ap-northeast-1"],
    "日本": ["ap-northeast-1"],
    "首尔": ["ap-northeast-2"],
    "韩国": ["ap-northeast-2"],
    "新加坡": ["ap-southeast-1"],
    "悉尼": ["ap-southeast-2"],
    "澳洲": ["ap-southeast-2"],
    "孟买": ["ap-south-1"],
    "印度": ["ap-south-1"],
    "圣保罗": ["sa-east-1"],
    "巴西": ["sa-east-1"],
    "加拿大": ["ca-central-1"],
    "香港": ["ap-east-1"],
}

# 模型简写 → 完整 model ID 映射
MODEL_ALIAS = {
    "4.5": "anthropic.claude-sonnet-4-5-20250929-v1:0",
    "sonnet 4.5": "anthropic.claude-sonnet-4-5-20250929-v1:0",
    "sonnet4.5": "anthropic.claude-sonnet-4-5-20250929-v1:0",
    "4.6": "anthropic.claude-sonnet-4-6",
    "sonnet 4.6": "anthropic.claude-sonnet-4-6",
    "sonnet4.6": "anthropic.claude-sonnet-4-6",
    "4.7": "anthropic.claude-opus-4-7",
    "opus 4.7": "anthropic.claude-opus-4-7",
    "opus4.7": "anthropic.claude-opus-4-7",
    "4.8": "anthropic.claude-opus-4-8",
    "opus 4.8": "anthropic.claude-opus-4-8",
    "opus4.8": "anthropic.claude-opus-4-8",
    "fable 5": "anthropic.claude-fable-5",
    "fable5": "anthropic.claude-fable-5",
    "5": "anthropic.claude-fable-5",
    "opus 4.5": "anthropic.claude-opus-4-5-20251101-v1:0",
    "opus4.5": "anthropic.claude-opus-4-5-20251101-v1:0",
    "opus 4.6": "anthropic.claude-opus-4-6-v1",
    "opus4.6": "anthropic.claude-opus-4-6-v1",
    "haiku 4.5": "anthropic.claude-haiku-4-5-20251001-v1:0",
    "haiku4.5": "anthropic.claude-haiku-4-5-20251001-v1:0",
    "opus 4": "anthropic.claude-opus-4-20250514-v1:0",
    "sonnet 4": "anthropic.claude-sonnet-4-20250514-v1:0",
}

# 完整模型信息（同步自 app.py）
CLAUDE_VERSIONS = [
    {
        "id": "anthropic.claude-fable-5",
        "label": "Claude Fable 5",
        "sources": {
            "us": "us.anthropic.claude-fable-5",
            "eu": "eu.anthropic.claude-fable-5",
            "global": "global.anthropic.claude-fable-5",
        },
    },
    {
        "id": "anthropic.claude-opus-4-8",
        "label": "Claude Opus 4.8",
        "sources": {"global": "global.anthropic.claude-opus-4-8"},
    },
    {
        "id": "anthropic.claude-opus-4-7",
        "label": "Claude Opus 4.7",
        "sources": {"global": "global.anthropic.claude-opus-4-7"},
    },
    {
        "id": "anthropic.claude-opus-4-6-v1",
        "label": "Claude Opus 4.6",
        "sources": {
            "us": "us.anthropic.claude-opus-4-6-v1",
            "eu": "eu.anthropic.claude-opus-4-6-v1",
            "global": "global.anthropic.claude-opus-4-6-v1",
        },
    },
    {
        "id": "anthropic.claude-opus-4-5-20251101-v1:0",
        "label": "Claude Opus 4.5",
        "sources": {"global": "global.anthropic.claude-opus-4-5-20251101-v1:0"},
    },
    {
        "id": "anthropic.claude-sonnet-4-6",
        "label": "Claude Sonnet 4.6",
        "sources": {
            "us": "us.anthropic.claude-sonnet-4-6",
            "eu": "eu.anthropic.claude-sonnet-4-6",
            "global": "global.anthropic.claude-sonnet-4-6",
        },
    },
    {
        "id": "anthropic.claude-sonnet-4-5-20250929-v1:0",
        "label": "Claude Sonnet 4.5",
        "sources": {
            "us": "us.anthropic.claude-sonnet-4-5-20250929-v1:0",
            "eu": "eu.anthropic.claude-sonnet-4-5-20250929-v1:0",
            "global": "global.anthropic.claude-sonnet-4-5-20250929-v1:0",
        },
    },
    {
        "id": "anthropic.claude-haiku-4-5-20251001-v1:0",
        "label": "Claude Haiku 4.5",
        "sources": {
            "us": "us.anthropic.claude-haiku-4-5-20251001-v1:0",
            "eu": "eu.anthropic.claude-haiku-4-5-20251001-v1:0",
            "global": "global.anthropic.claude-haiku-4-5-20251001-v1:0",
        },
    },
    {
        "id": "anthropic.claude-opus-4-20250514-v1:0",
        "label": "Claude Opus 4",
        "sources": {
            "us": "us.anthropic.claude-opus-4-20250514-v1:0",
            "eu": "eu.anthropic.claude-opus-4-20250514-v1:0",
            "global": "global.anthropic.claude-opus-4-20250514-v1:0",
        },
    },
    {
        "id": "anthropic.claude-sonnet-4-20250514-v1:0",
        "label": "Claude Sonnet 4",
        "sources": {
            "us": "us.anthropic.claude-sonnet-4-20250514-v1:0",
            "eu": "eu.anthropic.claude-sonnet-4-20250514-v1:0",
            "global": "global.anthropic.claude-sonnet-4-20250514-v1:0",
        },
    },
]

CLAUDE_BY_ID = {v["id"]: v for v in CLAUDE_VERSIONS}

US_GEO_REGIONS = {"us-east-1", "us-east-2", "us-west-1", "us-west-2", "ca-central-1", "sa-east-1"}
EU_GEO_REGIONS = {
    "eu-west-1", "eu-west-2", "eu-west-3", "eu-central-1", "eu-central-2",
    "eu-north-1", "eu-south-1", "eu-south-2", "il-central-1",
    "me-central-1", "me-south-1", "af-south-1",
}


# ═══════════════════════════════════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════════════════════════════════

def resolve_regions(raw_input):
    """将中文简写/英文 region code 解析为 region 列表"""
    regions = []
    parts = re.split(r'[,，、\s]+', raw_input.strip())
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if re.match(r'^[a-z]{2}-[a-z]+-\d$', part):
            regions.append(part)
        elif part in REGION_ALIAS:
            regions.extend(REGION_ALIAS[part])
        else:
            found = False
            for alias, codes in REGION_ALIAS.items():
                if part in alias or alias in part:
                    regions.extend(codes)
                    found = True
                    break
            if not found:
                raise ValueError(f"无法识别的区域: '{part}'")
    return list(dict.fromkeys(regions))


def resolve_models(raw_input):
    """将模型简写解析为 model_id 列表"""
    models = []
    parts = re.split(r'[,，、]+', raw_input.strip())
    for part in parts:
        part = part.strip().lower()
        if not part:
            continue
        if part.startswith("anthropic."):
            if part in CLAUDE_BY_ID:
                models.append(part)
            else:
                raise ValueError(f"未知模型 ID: '{part}'")
        elif part in MODEL_ALIAS:
            models.append(MODEL_ALIAS[part])
        else:
            found = False
            for alias, mid in MODEL_ALIAS.items():
                if part in alias or alias in part:
                    models.append(mid)
                    found = True
                    break
            if not found:
                raise ValueError(f"无法识别的模型: '{part}'")
    return list(dict.fromkeys(models))


def _region_geo(region):
    if region in US_GEO_REGIONS:
        return "us"
    if region in EU_GEO_REGIONS:
        return "eu"
    return "global"


def _inference_profile_arn(region, profile_id):
    return f"arn:aws:bedrock:{region}::inference-profile/{profile_id}"


def _resolve_copy_from(br, region, ver):
    """解析 create_inference_profile 的 copyFrom 来源 ARN"""
    sources = ver.get("sources") or {}
    geo = _region_geo(region)
    for key in (geo, "global"):
        pid = sources.get(key)
        if pid:
            return _inference_profile_arn(region, pid)
    return None


def _model_slug(model_id):
    return model_id.replace("anthropic.", "").replace(":", "-").replace(".", "-")[:36]


def _aws_error(e):
    if hasattr(e, 'response'):
        msg = e.response.get('Error', {}).get('Message', str(e))
        return msg
    return str(e)


# ═══════════════════════════════════════════════════════════════════
# 核心自动化逻辑 —— 创建
# ═══════════════════════════════════════════════════════════════════

def auto_provision(ak, sk, link_id, models_raw, regions_raw, tag_key, tag_value,
                   profile_prefix=None, test_invoke=True):
    """
    全流程自动化:
      1. 验证账号
      2. 解析区域和模型
      3. 为每个 区域×模型 创建 Inference Profile
      4. 给所有 Profile 打上标签
      5. 测试每个 Profile 是否可用
      6. 返回结果列表
    """
    results = []

    # Step 1: 验证账号
    try:
        sess = boto3.Session(aws_access_key_id=ak, aws_secret_access_key=sk)
        sts = sess.client("sts")
        identity = sts.get_caller_identity()
        account_id = identity.get("Account", "")
    except Exception as e:
        return {"ok": False, "error": f"账号验证失败: {_aws_error(e)}"}

    # 如果 link_id 为 auto，用 account_id
    if not link_id or link_id.lower() == "auto":
        link_id = account_id

    # Step 2: 解析输入
    try:
        regions = resolve_regions(regions_raw)
    except ValueError as e:
        return {"ok": False, "error": str(e)}

    try:
        model_ids = resolve_models(models_raw)
    except ValueError as e:
        return {"ok": False, "error": str(e)}

    if not regions:
        return {"ok": False, "error": "未指定有效区域"}
    if not model_ids:
        return {"ok": False, "error": "未指定有效模型"}

    prefix = profile_prefix or f"link-{link_id}"
    multi = len(regions) * len(model_ids) > 1

    # Step 3 & 4: 创建 Profile + 打标签
    for region in regions:
        br = sess.client("bedrock", region_name=region)
        br_runtime = sess.client("bedrock-runtime", region_name=region)

        for model_id in model_ids:
            ver = CLAUDE_BY_ID.get(model_id)
            if not ver:
                results.append({
                    "region": region, "model_id": model_id, "model_label": model_id,
                    "step": "resolve", "ok": False, "error": f"未知模型: {model_id}",
                })
                continue

            model_label = ver["label"]
            slug = _model_slug(model_id)
            profile_name = f"{prefix}-{region}-{slug}"[:64] if multi else prefix

            result_entry = {
                "region": region, "model_id": model_id, "model_label": model_label,
                "profile_name": profile_name, "profile_arn": "",
                "created": False, "tagged": False,
                "test_ok": None, "test_msg": "", "error": "",
            }

            copy_from_arn = _resolve_copy_from(br, region, ver)
            if not copy_from_arn:
                result_entry["error"] = "该区域无法找到可用的系统 Inference Profile"
                results.append(result_entry)
                continue

            # 创建
            try:
                tags = [{"key": tag_key, "value": tag_value}]
                resp = br.create_inference_profile(
                    inferenceProfileName=profile_name,
                    modelSource={"copyFrom": copy_from_arn},
                    description=f"Auto-provisioned for link {link_id}",
                    tags=tags,
                )
                profile_arn = resp.get("inferenceProfileArn", "")
                result_entry["profile_arn"] = profile_arn
                result_entry["created"] = True
                result_entry["tagged"] = True
            except botocore.exceptions.ClientError as e:
                code = e.response.get("Error", {}).get("Code", "")
                if code == "ConflictException" or "already exists" in str(e).lower():
                    try:
                        paginator = br.get_paginator("list_inference_profiles")
                        found_arn = None
                        for page in paginator.paginate(typeEquals="APPLICATION"):
                            for p in page.get("inferenceProfileSummaries", []):
                                if p.get("inferenceProfileName") == profile_name:
                                    found_arn = p.get("inferenceProfileArn", "")
                                    break
                            if found_arn:
                                break
                        if found_arn:
                            result_entry["profile_arn"] = found_arn
                            result_entry["created"] = True
                            result_entry["error"] = "已存在(跳过创建)"
                            try:
                                br.tag_resource(
                                    resourceARN=found_arn,
                                    tags=[{"key": tag_key, "value": tag_value}],
                                )
                                result_entry["tagged"] = True
                            except Exception as te:
                                result_entry["tagged"] = False
                                result_entry["error"] += f"; 打标签失败: {_aws_error(te)}"
                        else:
                            result_entry["error"] = "Profile 已存在但无法定位 ARN"
                    except Exception:
                        result_entry["error"] = "Profile 已存在，重新查找失败"
                else:
                    result_entry["error"] = _aws_error(e)
                    results.append(result_entry)
                    continue
            except Exception as e:
                result_entry["error"] = _aws_error(e)
                results.append(result_entry)
                continue

            # Step 5: 测试
            if test_invoke and result_entry["profile_arn"]:
                try:
                    test_body = json.dumps({
                        "anthropic_version": "bedrock-2023-05-31",
                        "max_tokens": 10,
                        "messages": [{"role": "user", "content": "Hi"}],
                    })
                    resp = br_runtime.invoke_model(
                        modelId=result_entry["profile_arn"],
                        body=test_body,
                        contentType="application/json",
                        accept="application/json",
                    )
                    body = json.loads(resp["body"].read())
                    result_entry["test_ok"] = True
                    result_entry["test_msg"] = "✓ 可用"
                except Exception as e:
                    result_entry["test_ok"] = False
                    result_entry["test_msg"] = f"✗ {_aws_error(e)}"

            results.append(result_entry)

    # 汇总
    created_cnt = sum(1 for r in results if r.get("created"))
    tagged_cnt = sum(1 for r in results if r.get("tagged"))
    test_ok_cnt = sum(1 for r in results if r.get("test_ok") is True)
    total = len(results)

    return {
        "ok": True,
        "account_id": account_id,
        "link_id": link_id,
        "tag": {"key": tag_key, "value": tag_value},
        "regions": regions,
        "models": [CLAUDE_BY_ID[m]["label"] for m in model_ids if m in CLAUDE_BY_ID],
        "results": results,
        "summary": {
            "total": total, "created": created_cnt, "tagged": tagged_cnt,
            "test_ok": test_ok_cnt, "test_fail": total - test_ok_cnt if test_invoke else 0,
        },
    }


# ═══════════════════════════════════════════════════════════════════
# 核心自动化逻辑 —— 删除
# ═══════════════════════════════════════════════════════════════════

def auto_delete(ak, sk, link_id=None, regions_raw=None, models_raw=None,
                profile_name_filter=None, delete_all=False):
    """
    批量删除 Inference Profile:
      - 按 link_id 前缀匹配（删除该客户所有 Profile）
      - 按 regions 过滤
      - 按 models 过滤（Profile 名称中包含 model slug）
      - 按 profile_name_filter 精确/模糊匹配
      - delete_all=True 删除指定区域的所有 APPLICATION Profile

    Returns:
        dict with keys: ok, account_id, deleted[], failed[], summary
    """
    # 验证账号
    try:
        sess = boto3.Session(aws_access_key_id=ak, aws_secret_access_key=sk)
        sts = sess.client("sts")
        identity = sts.get_caller_identity()
        account_id = identity.get("Account", "")
    except Exception as e:
        return {"ok": False, "error": f"账号验证失败: {_aws_error(e)}"}

    if not link_id or link_id.lower() == "auto":
        link_id = account_id

    # 解析区域
    if regions_raw:
        try:
            regions = resolve_regions(regions_raw)
        except ValueError as e:
            return {"ok": False, "error": str(e)}
    else:
        # 默认搜索所有常用区域
        regions = ["us-east-1", "us-east-2", "us-west-2", "us-west-1",
                   "eu-west-1", "eu-central-1", "ap-northeast-1", "ap-southeast-1"]

    # 解析模型（用于名称匹配）
    model_slugs = []
    if models_raw:
        try:
            model_ids = resolve_models(models_raw)
            model_slugs = [_model_slug(m) for m in model_ids]
        except ValueError:
            pass

    prefix = f"link-{link_id}"
    deleted = []
    failed = []

    for region in regions:
        br = sess.client("bedrock", region_name=region)

        # 列出该区域所有 APPLICATION Profile
        try:
            paginator = br.get_paginator("list_inference_profiles")
            profiles = []
            for page in paginator.paginate(typeEquals="APPLICATION"):
                for p in page.get("inferenceProfileSummaries", []):
                    profiles.append(p)
        except Exception as e:
            failed.append({
                "region": region, "profile_name": "(list failed)",
                "error": _aws_error(e),
            })
            continue

        # 过滤要删除的
        for p in profiles:
            pname = p.get("inferenceProfileName", "")
            parn = p.get("inferenceProfileArn", "")

            should_delete = False

            if delete_all:
                should_delete = True
            elif profile_name_filter:
                # 精确或模糊匹配
                if profile_name_filter in pname:
                    should_delete = True
            else:
                # 按 link_id 前缀匹配
                if pname.startswith(prefix):
                    # 如果指定了模型，进一步过滤
                    if model_slugs:
                        for slug in model_slugs:
                            if slug in pname:
                                should_delete = True
                                break
                    else:
                        should_delete = True

            if not should_delete:
                continue

            # 删除
            try:
                br.delete_inference_profile(inferenceProfileIdentifier=parn)
                deleted.append({
                    "region": region,
                    "profile_name": pname,
                    "profile_arn": parn,
                    "ok": True,
                })
            except Exception as e:
                failed.append({
                    "region": region,
                    "profile_name": pname,
                    "profile_arn": parn,
                    "error": _aws_error(e),
                    "ok": False,
                })

    return {
        "ok": True,
        "account_id": account_id,
        "link_id": link_id,
        "regions": regions,
        "deleted": deleted,
        "failed": failed,
        "summary": {
            "total_found": len(deleted) + len(failed),
            "deleted": len(deleted),
            "failed": len(failed),
        },
    }


# ═══════════════════════════════════════════════════════════════════
# Excel 导出
# ═══════════════════════════════════════════════════════════════════

def export_excel(provision_result):
    """将结果导出为 Excel BytesIO"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Provision Results"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="232F3E")
    ok_fill = PatternFill("solid", fgColor="E6F4EA")
    err_fill = PatternFill("solid", fgColor="FDE8E8")
    border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )

    ws.append(["Link ID", provision_result.get("link_id", "")])
    ws.append(["Account ID", provision_result.get("account_id", "")])
    tag = provision_result.get("tag", {})
    ws.append(["Tag", f"{tag.get('key', '')} = {tag.get('value', '')}"])
    ws.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])

    headers = ["区域", "模型", "Profile 名称", "Profile ARN", "已创建", "已打标", "测试结果", "备注"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r in provision_result.get("results", []):
        row = [
            r.get("region", ""), r.get("model_label", ""),
            r.get("profile_name", ""), r.get("profile_arn", ""),
            "✓" if r.get("created") else "✗",
            "✓" if r.get("tagged") else "✗",
            r.get("test_msg", "未测试"), r.get("error", ""),
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if not (r.get("created") and r.get("tagged")):
                cell.fill = err_fill

    col_widths = [14, 18, 30, 60, 8, 8, 12, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_delete_excel(delete_result):
    """将删除结果导出为 Excel BytesIO"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Delete Results"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="232F3E")
    border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )

    ws.append(["Link ID", delete_result.get("link_id", "")])
    ws.append(["Account ID", delete_result.get("account_id", "")])
    ws.append(["操作", "删除 Inference Profile"])
    ws.append(["时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])

    headers = ["区域", "Profile 名称", "Profile ARN", "状态", "备注"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r in delete_result.get("deleted", []):
        ws.append([r.get("region", ""), r.get("profile_name", ""),
                   r.get("profile_arn", ""), "✓ 已删除", ""])

    for r in delete_result.get("failed", []):
        ws.append([r.get("region", ""), r.get("profile_name", ""),
                   r.get("profile_arn", ""), "✗ 失败", r.get("error", "")])

    col_widths = [14, 30, 60, 10, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════
# Flask Web UI
# ═══════════════════════════════════════════════════════════════════

def create_app():
    from flask import Flask, render_template_string, request, jsonify, send_file
    from flask_cors import CORS

    app = Flask(__name__)
    CORS(app)

    HTML_PAGE = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Bedrock 一键配置</title>
<style>
:root{
  --bg:#eef2f7;--surface:#fff;--border:#e2e8f0;--text:#0f172a;
  --t2:#64748b;--pr:#2563eb;--ph:#1d4ed8;--ok:#16a34a;--er:#dc2626;
  --aws:#ff9900;--aws-dark:#232f3e;--r:8px;
  --shadow:0 1px 3px rgba(15,23,42,.06),0 4px 16px rgba(15,23,42,.04);
}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
  background:var(--bg);color:var(--text);font-size:14px;line-height:1.5;min-height:100vh}

.header{background:linear-gradient(135deg,var(--aws-dark),#1a365d);padding:12px 24px;
  display:flex;align-items:center;gap:12px;color:#fff}
.header h1{font-size:18px;font-weight:600}
.header .badge{background:var(--aws);color:var(--aws-dark);padding:2px 8px;
  border-radius:4px;font-size:11px;font-weight:700}
.header .nav{display:flex;gap:4px;margin-left:auto}
.header .nav button{background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.2);
  color:#fff;padding:6px 14px;border-radius:5px;cursor:pointer;font-size:12px;transition:all .2s}
.header .nav button:hover,.header .nav button.active{background:rgba(255,255,255,.2)}

.container{max-width:800px;margin:24px auto;padding:0 16px}
.card{background:var(--surface);border:1px solid var(--border);border-radius:var(--r);
  box-shadow:var(--shadow);padding:20px;margin-bottom:16px}
.card h2{font-size:14px;font-weight:600;margin-bottom:12px;color:var(--t2)}

.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.form-grid.full{grid-template-columns:1fr}
label{display:block;font-size:12px;font-weight:500;color:var(--t2);margin-bottom:4px}
input,textarea{width:100%;padding:8px 12px;border:1px solid var(--border);border-radius:6px;
  font-size:13px;transition:border-color .2s}
input:focus,textarea:focus{outline:none;border-color:var(--pr)}
textarea{resize:vertical;min-height:60px}

.btn{padding:10px 24px;border:none;border-radius:6px;font-size:14px;font-weight:600;
  cursor:pointer;transition:all .2s;margin:4px}
.btn-primary{background:var(--pr);color:#fff}
.btn-primary:hover{background:var(--ph)}
.btn-primary:disabled{opacity:.5;cursor:not-allowed}
.btn-danger{background:var(--er);color:#fff}
.btn-danger:hover{background:#b91c1c}
.btn-danger:disabled{opacity:.5;cursor:not-allowed}

.help{font-size:11px;color:var(--t2);margin-top:4px}

.results{margin-top:16px}
.result-item{display:grid;grid-template-columns:100px 140px 1fr 60px 60px;gap:8px;
  padding:8px 12px;border-bottom:1px solid var(--border);font-size:12px;align-items:center}
.result-item:first-child{font-weight:600;background:#f8fafc}
.result-item-del{display:grid;grid-template-columns:100px 200px 1fr 80px;gap:8px;
  padding:8px 12px;border-bottom:1px solid var(--border);font-size:12px;align-items:center}
.result-item-del:first-child{font-weight:600;background:#f8fafc}
.tag-ok{color:var(--ok);font-weight:600}
.tag-err{color:var(--er);font-weight:600}

.summary{background:#f0f9ff;border:1px solid #bae6fd;border-radius:6px;padding:12px 16px;
  margin:12px 0;font-size:13px}
.summary b{color:var(--pr)}
.summary.del{background:#fef2f2;border-color:#fecaca}
.summary.del b{color:var(--er)}

.spinner{display:inline-block;width:16px;height:16px;border:2px solid var(--border);
  border-top-color:var(--pr);border-radius:50%;animation:spin .6s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}

#log{background:#1e293b;color:#e2e8f0;font-family:monospace;font-size:12px;
  padding:12px;border-radius:6px;max-height:300px;overflow-y:auto;white-space:pre-wrap;
  display:none;margin-top:12px}

.tab-content{display:none}
.tab-content.active{display:block}
</style>
</head>
<body>

<div class="header">
  <h1>⚡ Bedrock 一键配置</h1>
  <span class="badge">AUTO</span>
  <div class="nav">
    <button class="active" onclick="switchTab('create')">🚀 创建配置</button>
    <button onclick="switchTab('delete')">🗑️ 删除配置</button>
  </div>
</div>

<div class="container">
  <!-- 凭证（共用） -->
  <div class="card">
    <h2>🔑 账号凭证</h2>
    <div class="form-grid">
      <div><label>Access Key</label><input id="ak" type="password" placeholder="AKIA..."></div>
      <div><label>Secret Key</label><input id="sk" type="password" placeholder="****"></div>
    </div>
  </div>

  <!-- ═══ 创建 Tab ═══ -->
  <div id="tab-create" class="tab-content active">
    <div class="card">
      <h2>📋 创建配置</h2>
      <div class="form-grid">
        <div><label>Link ID / Payer ID</label><input id="linkId" placeholder="例: 621724235498（留空自动获取）"></div>
        <div><label>区域</label><input id="regions" placeholder="例: 美西、美东">
          <div class="help">支持: 美东/美西/欧洲/东京/新加坡/香港... 或直接写 us-east-1</div>
        </div>
      </div>
      <div class="form-grid full" style="margin-top:12px">
        <div><label>模型</label><input id="models" placeholder="例: 4.5, 4.6, 4.7, 4.8, fable 5">
          <div class="help">支持简写: 4.5/4.6/4.7/4.8/fable 5/opus 4.5/haiku 4.5...</div>
        </div>
      </div>
      <div class="form-grid" style="margin-top:12px">
        <div><label>Tag Key</label><input id="tagKey" placeholder="例: migrationId"></div>
        <div><label>Tag Value</label><input id="tagValue" placeholder="例: migWEII3463IB"></div>
      </div>
    </div>
    <div class="card" style="text-align:center">
      <label style="margin-bottom:12px;display:flex;align-items:center;justify-content:center;gap:6px">
        <input type="checkbox" id="doTest" checked> 创建后测试可用性
      </label>
      <button class="btn btn-primary" id="runBtn" onclick="runCreate()">🚀 一键创建+打标</button>
      <div id="log"></div>
    </div>
    <div class="card" id="resultCard" style="display:none">
      <h2>📊 执行结果</h2>
      <div class="summary" id="summaryBox"></div>
      <div class="results" id="resultsGrid"></div>
      <div style="margin-top:16px;text-align:center">
        <button class="btn btn-primary" onclick="downloadExcel()">📥 下载 Excel</button>
      </div>
    </div>
  </div>

  <!-- ═══ 删除 Tab ═══ -->
  <div id="tab-delete" class="tab-content">
    <div class="card">
      <h2>🗑️ 删除配置</h2>
      <div class="form-grid">
        <div><label>Link ID（按前缀匹配删除）</label><input id="delLinkId" placeholder="例: 621724235498">
          <div class="help">将删除名称以 link-{ID} 开头的所有 Profile</div>
        </div>
        <div><label>区域</label><input id="delRegions" placeholder="例: 美西、美东（留空=搜索所有常用区域）"></div>
      </div>
      <div class="form-grid" style="margin-top:12px">
        <div><label>模型过滤（可选）</label><input id="delModels" placeholder="留空=删除该 Link 下所有模型">
          <div class="help">指定后只删除包含对应模型的 Profile</div>
        </div>
        <div><label>名称关键字过滤（可选）</label><input id="delFilter" placeholder="留空=按 Link ID 前缀匹配"></div>
      </div>
    </div>
    <div class="card" style="text-align:center">
      <button class="btn btn-danger" id="delBtn" onclick="runDelete()">🗑️ 一键删除</button>
      <div id="delLog" style="background:#1e293b;color:#e2e8f0;font-family:monospace;font-size:12px;
        padding:12px;border-radius:6px;max-height:300px;overflow-y:auto;white-space:pre-wrap;
        display:none;margin-top:12px"></div>
    </div>
    <div class="card" id="delResultCard" style="display:none">
      <h2>📊 删除结果</h2>
      <div class="summary del" id="delSummaryBox"></div>
      <div class="results" id="delResultsGrid"></div>
      <div style="margin-top:16px;text-align:center">
        <button class="btn btn-primary" onclick="downloadDeleteExcel()">📥 下载 Excel</button>
      </div>
    </div>
  </div>
</div>

<script>
let lastResult = null;
let lastDeleteResult = null;

function switchTab(tab) {
  document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
  document.getElementById('tab-' + tab).classList.add('active');
  document.querySelectorAll('.header .nav button').forEach(btn => btn.classList.remove('active'));
  event.target.classList.add('active');
}

function log(id, msg) {
  const el = document.getElementById(id);
  el.style.display = 'block';
  el.textContent += msg + '\n';
  el.scrollTop = el.scrollHeight;
}

// ═══ 创建 ═══
async function runCreate() {
  const btn = document.getElementById('runBtn');
  btn.disabled = true;
  btn.innerHTML = '<span class="spinner"></span> 执行中...';
  document.getElementById('log').textContent = '';
  document.getElementById('log').style.display = 'block';
  document.getElementById('resultCard').style.display = 'none';

  const payload = {
    access_key: document.getElementById('ak').value.trim(),
    secret_key: document.getElementById('sk').value.trim(),
    link_id: document.getElementById('linkId').value.trim() || 'auto',
    regions: document.getElementById('regions').value.trim(),
    models: document.getElementById('models').value.trim(),
    tag_key: document.getElementById('tagKey').value.trim(),
    tag_value: document.getElementById('tagValue').value.trim(),
    test_invoke: document.getElementById('doTest').checked,
  };

  if (!payload.access_key || !payload.secret_key) { alert('请填写 AK/SK'); btn.disabled=false; btn.textContent='🚀 一键创建+打标'; return; }
  if (!payload.regions) { alert('请填写区域'); btn.disabled=false; btn.textContent='🚀 一键创建+打标'; return; }
  if (!payload.models) { alert('请填写模型'); btn.disabled=false; btn.textContent='🚀 一键创建+打标'; return; }
  if (!payload.tag_key || !payload.tag_value) { alert('请填写 Tag Key 和 Value'); btn.disabled=false; btn.textContent='🚀 一键创建+打标'; return; }

  log('log', '▶ 开始执行...');
  log('log', `  Link ID: ${payload.link_id}`);
  log('log', `  区域: ${payload.regions}`);
  log('log', `  模型: ${payload.models}`);
  log('log', `  Tag: ${payload.tag_key} = ${payload.tag_value}`);

  try {
    const resp = await fetch('/api/auto_provision', { method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(payload) });
    const data = await resp.json();
    lastResult = data;
    if (!data.ok) { log('log', '\n✗ 失败: ' + data.error); }
    else {
      log('log', '\n✓ 执行完成!');
      log('log', `  账号: ${data.account_id}`);
      log('log', `  总计: ${data.summary.total} | 创建: ${data.summary.created} | 打标: ${data.summary.tagged} | 测试通过: ${data.summary.test_ok}`);
      showCreateResults(data);
    }
  } catch(e) { log('log', '\n✗ 请求失败: ' + e.message); }
  btn.disabled = false;
  btn.textContent = '🚀 一键创建+打标';
}

function showCreateResults(data) {
  document.getElementById('resultCard').style.display = 'block';
  const s = data.summary;
  document.getElementById('summaryBox').innerHTML =
    `<b>总计 ${s.total}</b> 个 Profile | 创建成功: <b>${s.created}</b> | 打标成功: <b>${s.tagged}</b> | 测试通过: <b>${s.test_ok}</b>`;
  let html = '<div class="result-item"><span>区域</span><span>模型</span><span>Profile ARN</span><span>标签</span><span>测试</span></div>';
  for (const r of data.results) {
    const tagCls = r.tagged ? 'tag-ok' : 'tag-err';
    const testCls = r.test_ok === true ? 'tag-ok' : (r.test_ok === false ? 'tag-err' : '');
    const testText = r.test_ok === true ? '✓' : (r.test_ok === false ? '✗' : '-');
    html += `<div class="result-item"><span>${r.region}</span><span>${r.model_label}</span>
      <span style="font-size:11px;word-break:break-all">${r.profile_arn || r.error}</span>
      <span class="${tagCls}">${r.tagged ? '✓' : '✗'}</span><span class="${testCls}">${testText}</span></div>`;
  }
  document.getElementById('resultsGrid').innerHTML = html;
}

async function downloadExcel() {
  if (!lastResult) { alert('没有结果可导出'); return; }
  const resp = await fetch('/api/auto_provision/excel', { method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(lastResult) });
  const blob = await resp.blob();
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a'); a.href = url;
  a.download = `provision_${lastResult.link_id}_${new Date().toISOString().slice(0,10)}.xlsx`;
  a.click(); URL.revokeObjectURL(url);
}

// ═══ 删除 ═══
async function runDelete() {
  const btn = document.getElementById('delBtn');
  const linkId = document.getElementById('delLinkId').value.trim();
  const filter = document.getElementById('delFilter').value.trim();

  if (!linkId && !filter) { alert('请填写 Link ID 或名称关键字'); return; }

  if (!confirm(`确定要删除 link-${linkId || filter} 相关的所有 Inference Profile 吗？此操作不可撤销！`)) return;

  btn.disabled = true;
  btn.innerHTML = '<span class="spinner"></span> 删除中...';
  document.getElementById('delLog').textContent = '';
  document.getElementById('delLog').style.display = 'block';
  document.getElementById('delResultCard').style.display = 'none';

  const payload = {
    access_key: document.getElementById('ak').value.trim(),
    secret_key: document.getElementById('sk').value.trim(),
    link_id: linkId || 'auto',
    regions: document.getElementById('delRegions').value.trim(),
    models: document.getElementById('delModels').value.trim(),
    profile_name_filter: filter,
  };

  if (!payload.access_key || !payload.secret_key) { alert('请填写 AK/SK'); btn.disabled=false; btn.textContent='🗑️ 一键删除'; return; }

  log('delLog', '▶ 开始删除...');
  log('delLog', `  Link ID: ${payload.link_id}`);
  log('delLog', `  区域: ${payload.regions || '所有常用区域'}`);

  try {
    const resp = await fetch('/api/auto_delete', { method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(payload) });
    const data = await resp.json();
    lastDeleteResult = data;
    if (!data.ok) { log('delLog', '\n✗ 失败: ' + data.error); }
    else {
      log('delLog', '\n✓ 删除完成!');
      log('delLog', `  删除: ${data.summary.deleted} | 失败: ${data.summary.failed}`);
      showDeleteResults(data);
    }
  } catch(e) { log('delLog', '\n✗ 请求失败: ' + e.message); }
  btn.disabled = false;
  btn.textContent = '🗑️ 一键删除';
}

function showDeleteResults(data) {
  document.getElementById('delResultCard').style.display = 'block';
  const s = data.summary;
  document.getElementById('delSummaryBox').innerHTML =
    `找到 <b>${s.total_found}</b> 个 Profile | 已删除: <b>${s.deleted}</b> | 失败: <b>${s.failed}</b>`;
  let html = '<div class="result-item-del"><span>区域</span><span>Profile 名称</span><span>ARN</span><span>状态</span></div>';
  for (const r of data.deleted) {
    html += `<div class="result-item-del"><span>${r.region}</span><span>${r.profile_name}</span>
      <span style="font-size:11px;word-break:break-all">${r.profile_arn||''}</span><span class="tag-ok">✓ 已删除</span></div>`;
  }
  for (const r of data.failed) {
    html += `<div class="result-item-del"><span>${r.region}</span><span>${r.profile_name}</span>
      <span style="font-size:11px;word-break:break-all">${r.profile_arn||''}</span><span class="tag-err">✗ ${r.error||''}</span></div>`;
  }
  document.getElementById('delResultsGrid').innerHTML = html;
}

async function downloadDeleteExcel() {
  if (!lastDeleteResult) { alert('没有结果可导出'); return; }
  const resp = await fetch('/api/auto_delete/excel', { method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(lastDeleteResult) });
  const blob = await resp.blob();
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a'); a.href = url;
  a.download = `delete_${lastDeleteResult.link_id}_${new Date().toISOString().slice(0,10)}.xlsx`;
  a.click(); URL.revokeObjectURL(url);
}
</script>
</body>
</html>"""

    @app.route("/")
    def index():
        return render_template_string(HTML_PAGE)

    # ── 创建 API ──
    @app.route("/api/auto_provision", methods=["POST"])
    def api_auto_provision():
        data = request.get_json() or {}
        ak = (data.get("access_key") or "").strip()
        sk = (data.get("secret_key") or "").strip()
        link_id = (data.get("link_id") or "auto").strip()
        models_raw = (data.get("models") or "").strip()
        regions_raw = (data.get("regions") or "").strip()
        tag_key = (data.get("tag_key") or "").strip()
        tag_value = (data.get("tag_value") or "").strip()
        test_invoke = data.get("test_invoke", True)

        if not all([ak, sk, models_raw, regions_raw, tag_key, tag_value]):
            return jsonify({"ok": False, "error": "参数不完整，请填写所有必填字段"}), 400

        result = auto_provision(
            ak=ak, sk=sk, link_id=link_id,
            models_raw=models_raw, regions_raw=regions_raw,
            tag_key=tag_key, tag_value=tag_value,
            test_invoke=test_invoke,
        )
        return jsonify(result)

    @app.route("/api/auto_provision/excel", methods=["POST"])
    def api_export_excel():
        data = request.get_json() or {}
        buf = export_excel(data)
        return send_file(
            buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"provision_{data.get('link_id', 'result')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        )

    # ── 删除 API ──
    @app.route("/api/auto_delete", methods=["POST"])
    def api_auto_delete():
        data = request.get_json() or {}
        ak = (data.get("access_key") or "").strip()
        sk = (data.get("secret_key") or "").strip()
        link_id = (data.get("link_id") or "").strip()
        regions_raw = (data.get("regions") or "").strip() or None
        models_raw = (data.get("models") or "").strip() or None
        profile_name_filter = (data.get("profile_name_filter") or "").strip() or None

        if not all([ak, sk]):
            return jsonify({"ok": False, "error": "请填写 AK/SK"}), 400
        if not link_id and not profile_name_filter:
            return jsonify({"ok": False, "error": "请填写 Link ID 或名称过滤关键字"}), 400

        result = auto_delete(
            ak=ak, sk=sk, link_id=link_id,
            regions_raw=regions_raw, models_raw=models_raw,
            profile_name_filter=profile_name_filter,
        )
        return jsonify(result)

    @app.route("/api/auto_delete/excel", methods=["POST"])
    def api_delete_excel():
        data = request.get_json() or {}
        buf = export_delete_excel(data)
        return send_file(
            buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"delete_{data.get('link_id', 'result')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        )

    return app


# ═══════════════════════════════════════════════════════════════════
# CLI 模式
# ═══════════════════════════════════════════════════════════════════

def run_cli_create(args):
    print(f"\n{'═'*60}")
    print(f"  Bedrock 一键配置 - 创建模式")
    print(f"{'═'*60}")
    print(f"  Link ID:   {args.link_id or 'auto'}")
    print(f"  模型:      {args.models}")
    print(f"  区域:      {args.regions}")
    print(f"  Tag:       {args.tag_key} = {args.tag_value}")
    print(f"{'═'*60}\n")

    result = auto_provision(
        ak=args.ak, sk=args.sk, link_id=args.link_id or "auto",
        models_raw=args.models, regions_raw=args.regions,
        tag_key=args.tag_key, tag_value=args.tag_value,
        test_invoke=not args.no_test,
    )

    if not result["ok"]:
        print(f"✗ 失败: {result['error']}")
        sys.exit(1)

    print(f"✓ 账号: {result['account_id']}")
    print(f"\n{'─'*60}")
    print(f"{'区域':<14} {'模型':<18} {'创建':<6} {'标签':<6} {'测试':<8} 备注")
    print(f"{'─'*60}")
    for r in result["results"]:
        created = "✓" if r.get("created") else "✗"
        tagged = "✓" if r.get("tagged") else "✗"
        test = r.get("test_msg", "-")
        error = r.get("error", "")
        print(f"{r['region']:<14} {r['model_label']:<18} {created:<6} {tagged:<6} {test:<8} {error}")

    s = result["summary"]
    print(f"\n{'─'*60}")
    print(f"汇总: 总计 {s['total']} | 创建 {s['created']} | 打标 {s['tagged']} | 测试通过 {s['test_ok']}")

    filename = args.output or f"provision_{result['link_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    buf = export_excel(result)
    with open(filename, "wb") as f:
        f.write(buf.read())
    print(f"\n📥 Excel 已保存: {filename}")


def run_cli_delete(args):
    print(f"\n{'═'*60}")
    print(f"  Bedrock 一键配置 - 删除模式")
    print(f"{'═'*60}")
    print(f"  Link ID:   {args.link_id or 'auto'}")
    print(f"  区域:      {args.regions or '所有常用区域'}")
    print(f"  模型过滤:  {args.models or '全部'}")
    print(f"{'═'*60}\n")

    result = auto_delete(
        ak=args.ak, sk=args.sk,
        link_id=args.link_id or "auto",
        regions_raw=args.regions or None,
        models_raw=args.models or None,
    )

    if not result["ok"]:
        print(f"✗ 失败: {result['error']}")
        sys.exit(1)

    print(f"✓ 账号: {result['account_id']}")
    print(f"\n{'─'*60}")
    print(f"{'区域':<14} {'Profile 名称':<40} 状态")
    print(f"{'─'*60}")
    for r in result["deleted"]:
        print(f"{r['region']:<14} {r['profile_name']:<40} ✓ 已删除")
    for r in result["failed"]:
        print(f"{r['region']:<14} {r['profile_name']:<40} ✗ {r.get('error','')}")

    s = result["summary"]
    print(f"\n{'─'*60}")
    print(f"汇总: 找到 {s['total_found']} | 已删除 {s['deleted']} | 失败 {s['failed']}")

    filename = args.output or f"delete_{result['link_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    buf = export_delete_excel(result)
    with open(filename, "wb") as f:
        f.write(buf.read())
    print(f"\n📥 Excel 已保存: {filename}")


# ═══════════════════════════════════════════════════════════════════
# Watch 守护模式 —— 监听 tasks/ 目录，自动执行任务
# ═══════════════════════════════════════════════════════════════════

def run_watch():
    """
    守护模式：监听 tasks/ 目录下的 .json 文件，自动执行并输出结果到 outputs/

    任务文件格式（创建）:
    {
        "action": "create",
        "ak": "...", "sk": "...",
        "link_id": "172229444780",
        "models": "4.5, 4.6, 4.7, 4.8, fable 5",
        "regions": "美东、美西",
        "tag_key": "map-migrated",
        "tag_value": "wjc-test"
    }

    任务文件格式（删除）:
    {
        "action": "delete",
        "ak": "...", "sk": "...",
        "link_id": "172229444780",
        "regions": "美东、美西"
    }
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    tasks_dir = os.path.join(base_dir, "tasks")
    outputs_dir = os.path.join(base_dir, "outputs")
    creds_file = os.path.join(base_dir, "creds.json")

    # 加载本地凭证文件（如果存在）
    default_creds = {}
    if os.path.exists(creds_file):
        try:
            with open(creds_file, "r", encoding="utf-8") as cf:
                default_creds = json.load(cf)
        except Exception:
            pass

    os.makedirs(tasks_dir, exist_ok=True)
    os.makedirs(outputs_dir, exist_ok=True)

    print(f"\n{'═'*60}")
    print(f"  👁️  Bedrock Watch 模式已启动")
    print(f"{'═'*60}")
    print(f"  监听目录: {tasks_dir}")
    print(f"  输出目录: {outputs_dir}")
    if default_creds.get("ak"):
        print(f"  凭证文件: {creds_file} ✓")
    print(f"  放入 .json 任务文件即可自动执行")
    print(f"  Ctrl+C 退出")
    print(f"{'═'*60}\n")

    while True:
        try:
            for fname in sorted(os.listdir(tasks_dir)):
                if not fname.endswith(".json"):
                    continue
                fpath = os.path.join(tasks_dir, fname)
                try:
                    with open(fpath, "r", encoding="utf-8") as f:
                        task = json.load(f)
                except Exception as e:
                    print(f"  ✗ 读取失败 {fname}: {e}")
                    os.rename(fpath, fpath + ".error")
                    continue

                action = task.get("action", "create")
                print(f"  ▶ 检测到任务: {fname} (action={action})")

                # 如果任务文件中没有有效 AK/SK，从 creds.json 读取
                if not task.get("ak") or "REDACTED" in task.get("ak", ""):
                    task["ak"] = default_creds.get("ak", "")
                if not task.get("sk") or "REDACTED" in task.get("sk", ""):
                    task["sk"] = default_creds.get("sk", "")

                if action == "create":
                    result = auto_provision(
                        ak=task.get("ak", ""), sk=task.get("sk", ""),
                        link_id=task.get("link_id", "auto"),
                        models_raw=task.get("models", ""),
                        regions_raw=task.get("regions", ""),
                        tag_key=task.get("tag_key", ""),
                        tag_value=task.get("tag_value", ""),
                        test_invoke=task.get("test_invoke", True),
                    )
                    # 保存 Excel
                    if result.get("ok"):
                        excel_name = f"provision_{result.get('link_id', 'result')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        buf = export_excel(result)
                        with open(os.path.join(outputs_dir, excel_name), "wb") as ef:
                            ef.write(buf.read())
                        print(f"  ✓ 完成! Excel: outputs/{excel_name}")
                    else:
                        print(f"  ✗ 失败: {result.get('error')}")

                elif action == "delete":
                    result = auto_delete(
                        ak=task.get("ak", ""), sk=task.get("sk", ""),
                        link_id=task.get("link_id", ""),
                        regions_raw=task.get("regions") or None,
                        models_raw=task.get("models") or None,
                    )
                    if result.get("ok"):
                        excel_name = f"delete_{result.get('link_id', 'result')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        buf = export_delete_excel(result)
                        with open(os.path.join(outputs_dir, excel_name), "wb") as ef:
                            ef.write(buf.read())
                        print(f"  ✓ 删除完成! Excel: outputs/{excel_name}")
                    else:
                        print(f"  ✗ 删除失败: {result.get('error')}")

                # 保存结果 JSON
                result_path = os.path.join(outputs_dir, fname.replace(".json", "_result.json"))
                with open(result_path, "w", encoding="utf-8") as rf:
                    json.dump(result, rf, ensure_ascii=False, indent=2)

                # 移除已处理的任务文件
                os.remove(fpath)

            time.sleep(2)  # 每 2 秒检查一次

        except KeyboardInterrupt:
            print("\n\n👋 Watch 模式已退出")
            break
        except Exception as e:
            print(f"  ⚠️ 异常: {e}")
            time.sleep(5)


# ═══════════════════════════════════════════════════════════════════
# 入口
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Bedrock Inference Profile 一键配置/删除")
    parser.add_argument("--cli", action="store_true", help="CLI 模式（否则启动 Web 服务）")
    parser.add_argument("--watch", action="store_true", help="Watch 守护模式（监听 tasks/ 目录自动执行）")
    parser.add_argument("--delete", action="store_true", help="删除模式（默认为创建模式）")
    parser.add_argument("--link-id", help="Link ID / Payer ID（留空自动获取）")
    parser.add_argument("--models", help="模型列表（逗号分隔）")
    parser.add_argument("--regions", help="区域列表（逗号分隔）")
    parser.add_argument("--tag-key", help="Tag Key")
    parser.add_argument("--tag-value", help="Tag Value")
    parser.add_argument("--ak", help="AWS Access Key")
    parser.add_argument("--sk", help="AWS Secret Key")
    parser.add_argument("--no-test", action="store_true", help="跳过测试")
    parser.add_argument("--output", "-o", help="Excel 输出路径")
    parser.add_argument("--port", type=int, default=5050, help="Web 端口（默认 5050）")

    args = parser.parse_args()

    if args.watch:
        run_watch()
    elif args.cli:
        if not all([args.ak, args.sk]):
            parser.error("CLI 模式需要至少提供 --ak 和 --sk")

        if args.delete:
            if not args.link_id and not args.models:
                parser.error("删除模式需要: --link-id 或 --models")
            run_cli_delete(args)
        else:
            if not all([args.models, args.regions, args.tag_key, args.tag_value]):
                parser.error("创建模式需要: --models, --regions, --tag-key, --tag-value")
            run_cli_create(args)
    else:
        app = create_app()
        print(f"\n🚀 Bedrock 一键配置工具已启动")
        print(f"   http://localhost:{args.port}")
        print(f"   Ctrl+C 退出\n")
        app.run(host="0.0.0.0", port=args.port, debug=True)
