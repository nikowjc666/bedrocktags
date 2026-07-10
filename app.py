"""AWS Bedrock Inference Profile 管理工具"""
import re
import json
import threading
import requests
import boto3
import botocore
from flask import Flask, render_template, request, jsonify, send_file, Response, stream_with_context
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime

app = Flask(__name__)
CORS(app)

REGIONS = [
    "us-east-1", "us-east-2", "us-west-2", "us-west-1",
    "eu-west-1", "eu-west-2", "eu-west-3", "eu-central-1",
    "eu-central-2", "eu-north-1", "eu-south-1", "eu-south-2",
    "ap-east-1", "ap-south-1", "ap-south-2",
    "ap-southeast-1", "ap-southeast-2", "ap-southeast-3", "ap-southeast-4",
    "ap-northeast-1", "ap-northeast-2", "ap-northeast-3",
    "sa-east-1", "ca-central-1", "me-south-1", "me-central-1",
    "af-south-1", "il-central-1",
]

# 固定 Claude 4+ 版本（无需每次从 AWS 加载）
# sources: 系统 Inference Profile ID，用于 copyFrom（新版模型不支持直接用 foundation model）
CLAUDE_VERSIONS = [
    {
        "id": "anthropic.claude-sonnet-5",
        "label": "Claude Sonnet 5",
        "sources": {
            "us": "us.anthropic.claude-sonnet-5",
            "global": "global.anthropic.claude-sonnet-5",
        },
    },
    {
        "id": "anthropic.claude-fable-5",
        "label": "Claude Fable 5",
        "sources": {
            "us": "us.anthropic.claude-fable-5",
            "eu": "eu.anthropic.claude-fable-5",
            "global": "global.anthropic.claude-fable-5",
        },
    },
    {
        "id": "anthropic.claude-opus-4-8",
        "label": "Claude Opus 4.8",
        "sources": {"global": "global.anthropic.claude-opus-4-8"},
    },
    {
        "id": "anthropic.claude-opus-4-7",
        "label": "Claude Opus 4.7",
        "sources": {"global": "global.anthropic.claude-opus-4-7"},
    },
    {
        "id": "anthropic.claude-opus-4-6-v1",
        "label": "Claude Opus 4.6",
        "sources": {
            "us": "us.anthropic.claude-opus-4-6-v1",
            "eu": "eu.anthropic.claude-opus-4-6-v1",
            "global": "global.anthropic.claude-opus-4-6-v1",
        },
    },
    {
        "id": "anthropic.claude-opus-4-5-20251101-v1:0",
        "label": "Claude Opus 4.5",
        "sources": {"global": "global.anthropic.claude-opus-4-5-20251101-v1:0"},
    },
    {
        "id": "anthropic.claude-sonnet-4-6",
        "label": "Claude Sonnet 4.6",
        "sources": {
            "us": "us.anthropic.claude-sonnet-4-6",
            "eu": "eu.anthropic.claude-sonnet-4-6",
            "global": "global.anthropic.claude-sonnet-4-6",
        },
    },
    {
        "id": "anthropic.claude-sonnet-4-5-20250929-v1:0",
        "label": "Claude Sonnet 4.5",
        "sources": {
            "us": "us.anthropic.claude-sonnet-4-5-20250929-v1:0",
            "eu": "eu.anthropic.claude-sonnet-4-5-20250929-v1:0",
            "global": "global.anthropic.claude-sonnet-4-5-20250929-v1:0",
        },
    },
    {
        "id": "anthropic.claude-haiku-4-5-20251001-v1:0",
        "label": "Claude Haiku 4.5",
        "sources": {
            "us": "us.anthropic.claude-haiku-4-5-20251001-v1:0",
            "eu": "eu.anthropic.claude-haiku-4-5-20251001-v1:0",
            "global": "global.anthropic.claude-haiku-4-5-20251001-v1:0",
        },
    },
    {
        "id": "anthropic.claude-opus-4-20250514-v1:0",
        "label": "Claude Opus 4",
        "sources": {
            "us": "us.anthropic.claude-opus-4-20250514-v1:0",
            "eu": "eu.anthropic.claude-opus-4-20250514-v1:0",
            "global": "global.anthropic.claude-opus-4-20250514-v1:0",
        },
    },
    {
        "id": "anthropic.claude-sonnet-4-20250514-v1:0",
        "label": "Claude Sonnet 4",
        "sources": {
            "us": "us.anthropic.claude-sonnet-4-20250514-v1:0",
            "eu": "eu.anthropic.claude-sonnet-4-20250514-v1:0",
            "global": "global.anthropic.claude-sonnet-4-20250514-v1:0",
        },
    },
]

CLAUDE_BY_ID = {v["id"]: v for v in CLAUDE_VERSIONS}

# Claude 4.5 及以上（排除 4.0 初代）
CLAUDE_40_IDS = {
    "anthropic.claude-opus-4-20250514-v1:0",
    "anthropic.claude-sonnet-4-20250514-v1:0",
}
CLAUDE_45_PLUS_VERSIONS = [v for v in CLAUDE_VERSIONS if v["id"] not in CLAUDE_40_IDS]
CLAUDE_45_BY_ID = {v["id"]: v for v in CLAUDE_45_PLUS_VERSIONS}


def _is_claude_45_plus(model_id):
    return model_id in CLAUDE_45_BY_ID


def _model_slug(model_id):
    return model_id.replace("anthropic.", "").replace(":", "-").replace(".", "-")[:36]


def _make_profile_name(prefix, region, model_id, multi):
    """生成 Inference Profile 名称。
    
    格式：claude-{版本}-auto{月日}-{前缀}
    例如：claude-sonnet-45-auto0615-myprofile
    """
    # 从 model_id 获取版本标签，去掉 "claude " 前缀避免重复
    ver = CLAUDE_45_BY_ID.get(model_id) or CLAUDE_BY_ID.get(model_id) or {}
    raw_label = (ver.get("label") or model_id)
    # 去掉开头的 "claude " / "Claude "（不区分大小写）
    raw_label = re.sub(r"(?i)^claude\s*", "", raw_label).strip()
    label = raw_label.lower().replace(" ", "-").replace(".", "")
    
    # 月份日期（如 0615）
    auto_date = datetime.now().strftime("%m%d")
    
    # 组装名称：claude-{版本}-auto{月日}-{前缀}
    parts = ["claude", label, f"auto{auto_date}"]
    if prefix:
        parts.append(prefix)
    name = "-".join(p for p in parts if p)
    return name[:64]

US_GEO_REGIONS = {
    "us-east-1", "us-east-2", "us-west-1", "us-west-2",
    "ca-central-1", "sa-east-1",
}
EU_GEO_REGIONS = {
    "eu-west-1", "eu-west-2", "eu-west-3", "eu-central-1", "eu-central-2",
    "eu-north-1", "eu-south-1", "eu-south-2", "il-central-1",
    "me-central-1", "me-south-1", "af-south-1",
}
AP_NORTHEAST = {"ap-northeast-1", "ap-northeast-2", "ap-northeast-3"}


def _foundation_model_arn(region, model_id):
    return f"arn:aws:bedrock:{region}::foundation-model/{model_id}"


def _inference_profile_arn(region, profile_id):
    return f"arn:aws:bedrock:{region}::inference-profile/{profile_id}"


def _region_geo(region):
    if region in US_GEO_REGIONS:
        return "us"
    if region in EU_GEO_REGIONS:
        return "eu"
    if region in AP_NORTHEAST:
        return "jp"
    return "global"


def _build_model_source_index(br, region):
    """从 AWS 列出系统 Inference Profile，建立 model_id -> copyFrom ARN 索引"""
    geo = _region_geo(region)
    index = {}
    paginator = br.get_paginator("list_inference_profiles")
    for page in paginator.paginate(typeEquals="SYSTEM_DEFINED"):
        for summary in page.get("inferenceProfileSummaries", []):
            pid = summary.get("inferenceProfileId", "")
            if "anthropic" not in pid:
                continue
            try:
                detail = br.get_inference_profile(inferenceProfileIdentifier=pid)
                profile = detail.get("inferenceProfile", detail)
                parn = profile.get("inferenceProfileArn", "")
                if not parn:
                    continue
                model_ids = set()
                in_region = False
                for m in profile.get("models", []):
                    arn = m.get("modelArn", "")
                    if f":bedrock:{region}:" in arn:
                        in_region = True
                    if "/foundation-model/" in arn:
                        model_ids.add(arn.split("/foundation-model/", 1)[1])
                priority = 0
                if pid.startswith(f"{geo}."):
                    priority = 3
                elif pid.startswith("global."):
                    priority = 2
                elif in_region:
                    priority = 1
                for mid in model_ids:
                    prev = index.get(mid)
                    if prev is None or priority > prev[0]:
                        index[mid] = (priority, parn)
            except Exception:
                continue
    return {k: v[1] for k, v in index.items()}


def _get_source_type(source_id):
    """根据源 ID 判断是否为 global 类型"""
    if not source_id:
        return "unknown"
    if source_id.startswith("global."):
        return "global"
    elif source_id.startswith("us."):
        return "us"
    elif source_id.startswith("eu."):
        return "eu"
    else:
        return "foundation"


def _resolve_copy_from(br, region, ver, index=None):
    """解析 create_inference_profile 的 copyFrom 来源 ARN"""
    model_id = ver["id"]
    sources = ver.get("sources") or {}
    geo = _region_geo(region)
    # 强制优先使用 global，然后才是地理区域
    for key in ("global", geo):
        pid = sources.get(key)
        if pid:
            return _inference_profile_arn(region, pid), pid

    if index is None:
        index = _build_model_source_index(br, region)
    if model_id in index:
        return index[model_id], model_id

    try:
        resp = br.list_foundation_models(byInferenceType="ON_DEMAND")
        for m in resp.get("modelSummaries", []):
            if m.get("modelId") == model_id:
                arn = m.get("modelArn") or _foundation_model_arn(region, model_id)
                return arn, model_id
    except Exception:
        pass

    return None, None


def _creds(data):
    ak = (data.get("access_key") or "").strip()
    sk = (data.get("secret_key") or "").strip()
    user_id = (data.get("user_id") or "").strip()
    return ak, sk, user_id


# ── boto3 客户端缓存（相同 ak/sk/region 复用连接，避免重复握手）──
_client_cache: dict = {}

def _get_client(service, ak, sk, region):
    key = (service, ak, sk, region)
    if key not in _client_cache:
        cfg = botocore.config.Config(
            max_pool_connections=20,        # 连接池大小（支持并发）
            connect_timeout=5,
            read_timeout=30,
            retries={"max_attempts": 2},
        )
        sess = boto3.Session(aws_access_key_id=ak, aws_secret_access_key=sk)
        _client_cache[key] = sess.client(service, region_name=region, config=cfg)
    return _client_cache[key]


def _bedrock(ak, sk, region):
    return _get_client("bedrock", ak, sk, region)


def _bedrock_runtime(ak, sk, region):
    return _get_client("bedrock-runtime", ak, sk, region)


def _list_resource_tags(br, resource_arn):
    try:
        resp = br.list_tags_for_resource(resourceARN=resource_arn)
        return resp.get("tags", [])
    except Exception:
        return []


def _is_claude(model_summary):
    mid = (model_summary.get("modelId") or "").lower()
    prov = (model_summary.get("providerName") or "").lower()
    name = (model_summary.get("modelName") or "").lower()
    return "claude" in mid or "anthropic" in prov or "claude" in name


def _claude_models(br):
    resp = br.list_foundation_models()
    models = []
    for m in resp.get("modelSummaries", []):
        if not _is_claude(m):
            continue
        models.append({
            "modelArn": m.get("modelArn", ""),
            "modelId": m.get("modelId", ""),
            "modelName": m.get("modelName", ""),
            "providerName": m.get("providerName", ""),
            "inputModalities": m.get("inputModalities", []),
            "outputModalities": m.get("outputModalities", []),
        })
    return models


def _verify_account(ak, sk, user_id=None):
    sess = boto3.Session(aws_access_key_id=ak, aws_secret_access_key=sk)
    sts = sess.client("sts")
    identity = sts.get_caller_identity()
    account = identity["Account"]
    arn = identity.get("Arn", "")
    if user_id and user_id != account:
        return None, f"用户 ID 不匹配，期望 {user_id}，实际账号 {account}"
    return {"account_id": account, "arn": arn}, None


def _aws_error(e):
    if isinstance(e, botocore.exceptions.ClientError):
        return f"AWS 错误: {e.response['Error']['Message']}"
    return str(e)


def _profile_summary(p, tags=None):
    summary = {
        "arn": p.get("inferenceProfileArn", ""),
        "name": p.get("inferenceProfileName", ""),
        "id": p.get("inferenceProfileId", ""),
        "status": p.get("status", ""),
        "description": p.get("description", ""),
        "type": p.get("type", ""),
        "createdAt": str(p.get("createdAt", "")),
        "updatedAt": str(p.get("updatedAt", "")),
    }
    if tags is not None:
        summary["tags"] = tags
    return summary


def _list_application_profiles(br, name_prefix=None):
    """列出 APPLICATION 类型 inference profile，可按名称关键词过滤（包含匹配）"""
    profiles = []
    paginator = br.get_paginator("list_inference_profiles")
    for page in paginator.paginate(typeEquals="APPLICATION"):
        for p in page.get("inferenceProfileSummaries", []):
            name = p.get("inferenceProfileName", "")
            if name_prefix and name_prefix.lower() not in name.lower():
                continue
            profiles.append(_profile_summary(p))
    return profiles



@app.route("/api/verify", methods=["POST"])
def verify():
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    if not ak or not sk:
        return jsonify({"ok": False, "error": "请填写 Access Key 和 Secret Key"}), 400
    try:
        info, err = _verify_account(ak, sk, user_id or None)
        if err:
            return jsonify({"ok": False, "error": err}), 400
        # 验证成功后后台预热 code_map
        def _prewarm():
            if not _quota_code_map:
                _build_code_map(ak, sk)
        threading.Thread(target=_prewarm, daemon=True).start()
        return jsonify({"ok": True, **info})
    except Exception as e:
        return jsonify({"ok": False, "error": _aws_error(e)}), 400


@app.route("/api/regions", methods=["GET"])
def regions():
    return jsonify({"ok": True, "regions": REGIONS})


@app.route("/api/claude_versions", methods=["GET"])
def claude_versions():
    return jsonify({"ok": True, "versions": CLAUDE_VERSIONS})


@app.route("/api/claude_45_versions", methods=["GET"])
def claude_45_versions():
    return jsonify({"ok": True, "versions": CLAUDE_45_PLUS_VERSIONS})


@app.route("/api/resolveclaude_models", methods=["POST"])
def resolve_claude_models():
    """解析 Claude 4.5+ 系统 Inference Profile ARN（区域 × 模型）"""
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    regions = [r.strip() for r in (data.get("regions") or []) if r and r.strip()]
    model_ids = [m.strip() for m in (data.get("model_ids") or []) if m and m.strip()]
    if not all([ak, sk]):
        return jsonify({"ok": False, "error": "请填写 AK/SK"}), 400
    if not regions:
        return jsonify({"ok": False, "error": "请至少选择一个区域"}), 400
    if not model_ids:
        return jsonify({"ok": False, "error": "请至少选择一个模型"}), 400
    invalid = [m for m in model_ids if not _is_claude_45_plus(m)]
    if invalid:
        return jsonify({"ok": False, "error": f"仅支持 Claude 4.5+，无效模型: {', '.join(invalid)}"}), 400
    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400
        results = []
        for region in regions:
            br = _bedrock(ak, sk, region)
            index = _build_model_source_index(br, region)
            for model_id in model_ids:
                ver = CLAUDE_45_BY_ID[model_id]
                arn, source_id = _resolve_copy_from(br, region, ver, index)
                if not arn:
                    results.append({
                        "region": region,
                        "model_id": model_id,
                        "model_label": ver["label"],
                        "ok": False,
                        "error": "该区域未找到可用的系统 Inference Profile",
                    })
                    continue
                try:
                    resp = br.get_inference_profile(inferenceProfileIdentifier=arn)
                    p = resp.get("inferenceProfile", resp)
                    results.append({
                        "region": region,
                        "model_id": model_id,
                        "model_label": ver["label"],
                        "profile_id": source_id or p.get("inferenceProfileId", ""),
                        "arn": p.get("inferenceProfileArn", arn),
                        "status": p.get("status", ""),
                        "type": p.get("type", ""),
                        "tags": _list_resource_tags(br, p.get("inferenceProfileArn", arn)),
                        "ok": True,
                    })
                except Exception as e:
                    results.append({
                        "region": region,
                        "model_id": model_id,
                        "model_label": ver["label"],
                        "arn": arn,
                        "ok": False,
                        "error": _aws_error(e),
                    })
        ok_cnt = sum(1 for r in results if r.get("ok"))
        return jsonify({
            "ok": True,
            "total": len(results),
            "success": ok_cnt,
            "failed": len(results) - ok_cnt,
            "results": results,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": _aws_error(e)}), 500


@app.route("/api/queryclaude_models", methods=["POST"])
def query_claude_models():
    """根据 AK/SK/用户 ID 查询账号可用的 Claude 模型"""
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    region = (data.get("region") or "us-east-1").strip()
    if not all([ak, sk, user_id]):
        return jsonify({"ok": False, "error": "请填写 AK、SK 和用户 ID（账号 ID）"}), 400
    try:
        info, err = _verify_account(ak, sk, user_id)
        if err:
            return jsonify({"ok": False, "error": err}), 400
        br = _bedrock(ak, sk, region)
        models = _claude_models(br)
        return jsonify({
            "ok": True,
            "account_id": info["account_id"],
            "region": region,
            "count": len(models),
            "models": models,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": _aws_error(e)}), 400


@app.route("/api/list_foundation_models", methods=["POST"])
def list_foundation_models():
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    region = (data.get("region") or "").strip()
    if not all([ak, sk, region]):
        return jsonify({"ok": False, "error": "参数不完整"}), 400
    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400
        br = _bedrock(ak, sk, region)
        return jsonify({"ok": True, "models": _claude_models(br)})
    except Exception as e:
        return jsonify({"ok": False, "error": _aws_error(e)}), 400


@app.route("/api/list_profiles", methods=["POST"])
def list_profiles():
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    region = (data.get("region") or "").strip()
    if not all([ak, sk, region]):
        return jsonify({"ok": False, "error": "参数不完整"}), 400
    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400
        br = _bedrock(ak, sk, region)
        profiles = []
        paginator = br.get_paginator("list_inference_profiles")
        for page in paginator.paginate(typeEquals="APPLICATION"):
            for p in page.get("inferenceProfileSummaries", []):
                profiles.append(_profile_summary(p))
        return jsonify({"ok": True, "profiles": profiles})
    except Exception as e:
        return jsonify({"ok": False, "error": _aws_error(e)}), 400


@app.route("/api/get_profile", methods=["POST"])
def get_profile():
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    region = (data.get("region") or "").strip()
    profile_id = (data.get("inference_profile_id") or data.get("inference_profile_arn") or "").strip()
    if not all([ak, sk, region, profile_id]):
        return jsonify({"ok": False, "error": "参数不完整"}), 400
    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400
        br = _bedrock(ak, sk, region)
        resp = br.get_inference_profile(inferenceProfileIdentifier=profile_id)
        p = resp.get("inferenceProfile", resp)
        tags = []
        try:
            tag_resp = br.list_tags_for_resource(resourceARN=p.get("inferenceProfileArn", profile_id))
            tags = tag_resp.get("tags", [])
        except Exception:
            pass
        return jsonify({
            "ok": True,
            "profile": _profile_summary(p),
            "modelSource": p.get("models", p.get("modelSource", {})),
            "tags": tags,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": _aws_error(e)}), 400


@app.route("/api/create_profile", methods=["POST"])
def create_profile():
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    region = (data.get("region") or "").strip()
    name = (data.get("name") or "").strip()
    description = (data.get("description") or "").strip()
    model_source_arn = (data.get("model_source_arn") or "").strip()
    model_id = (data.get("model_id") or "").strip()
    tags_raw = data.get("tags") or {}
    if not all([ak, sk, region, name]):
        return jsonify({"ok": False, "error": "请填写必要参数"}), 400
    if not model_source_arn and not model_id:
        return jsonify({"ok": False, "error": "请选择模型"}), 400
    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400
        br = _bedrock(ak, sk, region)
        source_id = None
        if not model_source_arn and model_id:
            ver = CLAUDE_BY_ID.get(model_id)
            if ver:
                model_source_arn, source_id = _resolve_copy_from(br, region, ver)
            else:
                model_source_arn = _foundation_model_arn(region, model_id)
        if not model_source_arn:
            return jsonify({
                "ok": False,
                "error": "无法解析模型来源，请确认该区域已开通该 Claude 模型",
            }), 400
        params = {
            "inferenceProfileName": name,
            "modelSource": {"copyFrom": model_source_arn},
        }
        if description:
            params["description"] = description
        if tags_raw:
            params["tags"] = [{"key": k, "value": v} for k, v in tags_raw.items()]
        resp = br.create_inference_profile(**params)
        source_type = _get_source_type(source_id)
        return jsonify({"ok": True, "result": {
            "inferenceProfileArn": resp.get("inferenceProfileArn", ""),
            "inferenceProfileId": resp.get("inferenceProfileId", ""),
            "status": resp.get("status", ""),
            "sourceType": source_type,
            "isGlobal": source_type == "global",
            "copyFrom": source_id or model_source_arn,
        }})
    except Exception as e:
        return jsonify({"ok": False, "error": _aws_error(e)}), 400


@app.route("/api/batch_create_profiles", methods=["POST"])
def batch_create_profiles():
    """区域 × Claude 4.5+ 模型 笛卡尔积批量创建 Inference Profile 并打标签"""
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    regions = [r.strip() for r in (data.get("regions") or []) if r and r.strip()]
    model_ids = [m.strip() for m in (data.get("model_ids") or []) if m and m.strip()]
    name_prefix = (data.get("name_prefix") or data.get("name") or "").strip()
    description = (data.get("description") or "").strip()
    tags_raw = data.get("tags") or {}
    claude_45_only = data.get("claude_45_only", True)

    if not all([ak, sk]):
        return jsonify({"ok": False, "error": "请填写 AK/SK"}), 400
    if not name_prefix:
        return jsonify({"ok": False, "error": "请填写配置名称"}), 400
    if not regions:
        return jsonify({"ok": False, "error": "请至少选择一个区域"}), 400
    if not model_ids:
        return jsonify({"ok": False, "error": "请至少选择一个 Claude 版本"}), 400

    if claude_45_only:
        invalid = [m for m in model_ids if not _is_claude_45_plus(m)]
        if invalid:
            return jsonify({
                "ok": False,
                "error": f"仅支持 Claude 4.5+，无效模型: {', '.join(invalid)}",
            }), 400

    total = len(regions) * len(model_ids)
    multi = len(regions) > 1 or len(model_ids) > 1
    tag_list = [{"key": k, "value": v} for k, v in tags_raw.items()] if tags_raw else []
    results = []

    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400

        for region in regions:
            br = _bedrock(ak, sk, region)
            source_index = _build_model_source_index(br, region)
            for model_id in model_ids:
                ver = (CLAUDE_45_BY_ID if claude_45_only else CLAUDE_BY_ID).get(model_id)
                if not ver:
                    results.append({
                        "region": region,
                        "model_id": model_id,
                        "name": "",
                        "ok": False,
                        "error": "未知或不支持的模型版本",
                    })
                    continue
                name = _make_profile_name(name_prefix, region, model_id, multi)
                arn, source_id = _resolve_copy_from(br, region, ver, source_index)
                if not arn:
                    results.append({
                        "region": region,
                        "model_id": model_id,
                        "model_label": ver["label"],
                        "name": name,
                        "ok": False,
                        "error": "该区域未找到可用的系统 Inference Profile",
                    })
                    continue
                try:
                    params = {
                        "inferenceProfileName": name,
                        "modelSource": {"copyFrom": arn},
                    }
                    if description:
                        params["description"] = description
                    if tag_list:
                        params["tags"] = tag_list
                    resp = br.create_inference_profile(**params)
                    profile_arn = resp.get("inferenceProfileArn", "")
                    source_type = _get_source_type(source_id)
                    results.append({
                        "region": region,
                        "model_id": model_id,
                        "model_label": ver["label"],
                        "name": name,
                        "ok": True,
                        "copyFrom": source_id or arn,
                        "sourceType": source_type,
                        "isGlobal": source_type == "global",
                        "inferenceProfileArn": profile_arn,
                        "inferenceProfileId": resp.get("inferenceProfileId", ""),
                        "status": resp.get("status", ""),
                        "tags": tags_raw,
                    })
                except Exception as e:
                    results.append({
                        "region": region,
                        "model_id": model_id,
                        "model_label": ver["label"],
                        "name": name,
                        "ok": False,
                        "error": _aws_error(e),
                    })

        ok_cnt = sum(1 for r in results if r["ok"])
        return jsonify({
            "ok": True,
            "total": total,
            "success": ok_cnt,
            "failed": total - ok_cnt,
            "results": results,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": _aws_error(e)}), 500


@app.route("/api/batch_create_stream", methods=["POST"])
def batch_create_stream():
    """流式批量创建 —— 区域间并发，每完成一个立即推送 SSE。

    优化：
    1. _resolve_copy_from 优先走硬编码 sources，绝大多数模型直接命中，
       完全跳过 _build_model_source_index 的 N 次 get_inference_profile 调用。
    2. 各区域用 ThreadPoolExecutor 并发执行，IO 等待不再串行叠加。
    3. fallback 时才懒加载 source_index，且同一区域只查一次。
    """
    import queue
    from concurrent.futures import ThreadPoolExecutor, as_completed

    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    regions        = [r.strip() for r in (data.get("regions")    or []) if r and r.strip()]
    model_ids      = [m.strip() for m in (data.get("model_ids")  or []) if m and m.strip()]
    name_prefix    = (data.get("name_prefix") or data.get("name") or "").strip()
    description    = (data.get("description") or "").strip()
    tags_raw       = data.get("tags") or {}
    claude_45_only = data.get("claude_45_only", True)

    def _err(msg):
        return Response(
            f"data: {json.dumps({'type':'error','error':msg}, ensure_ascii=False)}\n\n",
            mimetype="text/event-stream", status=400,
        )

    if not all([ak, sk]):   return _err("请填写 AK/SK")
    if not name_prefix:     return _err("请填写配置名称")
    if not regions:         return _err("请至少选择一个区域")
    if not model_ids:       return _err("请至少选择一个 Claude 版本")
    if claude_45_only:
        invalid = [m for m in model_ids if not _is_claude_45_plus(m)]
        if invalid:         return _err(f"仅支持 Claude 4.5+，无效模型: {', '.join(invalid)}")

    total    = len(regions) * len(model_ids)
    multi    = total > 1
    tag_list = [{"key": k, "value": v} for k, v in tags_raw.items()] if tags_raw else []
    model_lookup = CLAUDE_45_BY_ID if claude_45_only else CLAUDE_BY_ID

    def _create_region(region):
        """在一个区域里依次创建所有模型，返回 list[row_dict]。"""
        rows = []
        try:
            br = _bedrock(ak, sk, region)
        except Exception as e:
            for mid in model_ids:
                ver = model_lookup.get(mid) or {}
                rows.append({
                    "region": region, "model_id": mid,
                    "model_label": ver.get("label", mid),
                    "name": _make_profile_name(name_prefix, region, mid, multi),
                    "ok": False, "error": _aws_error(e),
                })
            return rows

        # 懒加载：只有 sources 命中失败时才查一次
        _index_cache = {}

        def _get_index():
            if "v" not in _index_cache:
                try:
                    _index_cache["v"] = _build_model_source_index(br, region)
                except Exception:
                    _index_cache["v"] = {}
            return _index_cache["v"]

        for model_id in model_ids:
            ver = model_lookup.get(model_id)
            if not ver:
                rows.append({
                    "region": region, "model_id": model_id, "model_label": model_id,
                    "name": "", "ok": False, "error": "未知或不支持的模型版本",
                })
                continue

            name = _make_profile_name(name_prefix, region, model_id, multi)

            # 先尝试 sources 直接命中（不触发 AWS API）
            sources = ver.get("sources") or {}
            geo     = _region_geo(region)
            arn = source_id = None
            for key in (geo, "global"):
                pid = sources.get(key)
                if pid:
                    arn       = _inference_profile_arn(region, pid)
                    source_id = pid
                    break

            # fallback：查 AWS 系统 profile 索引
            if not arn:
                idx = _get_index()
                if model_id in idx:
                    arn = source_id = idx[model_id]

            if not arn:
                rows.append({
                    "region": region, "model_id": model_id,
                    "model_label": ver["label"], "name": name,
                    "ok": False,
                    "error": "该区域未找到可用的系统 Inference Profile",
                })
                continue

            try:
                params = {"inferenceProfileName": name, "modelSource": {"copyFrom": arn}}
                if description:
                    params["description"] = description
                if tag_list:
                    params["tags"] = tag_list
                resp        = br.create_inference_profile(**params)
                profile_arn = resp.get("inferenceProfileArn", "")
                source_type = _get_source_type(source_id)
                rows.append({
                    "region": region, "model_id": model_id,
                    "model_label": ver["label"], "name": name,
                    "ok": True,
                    "copyFrom": source_id or arn,
                    "sourceType": source_type,
                    "isGlobal": source_type == "global",
                    "inferenceProfileArn": profile_arn,
                    "inferenceProfileId": resp.get("inferenceProfileId", ""),
                    "status": resp.get("status", ""),
                    "tags": tags_raw,
                })
            except Exception as e:
                rows.append({
                    "region": region, "model_id": model_id,
                    "model_label": ver["label"], "name": name,
                    "ok": False, "error": _aws_error(e),
                })
        return rows

    def _generate():
        yield f"data: {json.dumps({'type':'start','total':total}, ensure_ascii=False)}\n\n"

        try:
            if user_id:
                _, err = _verify_account(ak, sk, user_id)
                if err:
                    yield f"data: {json.dumps({'type':'error','error':err}, ensure_ascii=False)}\n\n"
                    return
        except Exception as e:
            yield f"data: {json.dumps({'type':'error','error':_aws_error(e)}, ensure_ascii=False)}\n\n"
            return

        done    = 0
        # 并发数：区域数，最多 10 个线程（避免连接数爆炸）
        workers = min(len(regions), 10)

        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = {pool.submit(_create_region, r): r for r in regions}
            for future in as_completed(futures):
                try:
                    rows = future.result()
                except Exception as e:
                    region = futures[future]
                    rows = [{"region": region, "model_id": mid,
                             "model_label": model_lookup.get(mid, {}).get("label", mid),
                             "name": _make_profile_name(name_prefix, region, mid, multi),
                             "ok": False, "error": _aws_error(e)}
                            for mid in model_ids]

                for row in rows:
                    done += 1
                    row.update({"type": "item", "done": done, "total": total})
                    yield f"data: {json.dumps(row, ensure_ascii=False)}\n\n"

        yield f"data: {json.dumps({'type':'done','total':total,'done':done}, ensure_ascii=False)}\n\n"

    return Response(
        stream_with_context(_generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/update_profile", methods=["POST"])
def update_profile():
    """更新配置标签（Bedrock 不支持直接修改名称/模型，需删除后重建）"""
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    region = (data.get("region") or "").strip()
    profile_arn = (data.get("inference_profile_arn") or "").strip()
    tags_raw = data.get("tags") or {}
    remove_tag_keys = data.get("remove_tag_keys") or []
    if not all([ak, sk, region, profile_arn]):
        return jsonify({"ok": False, "error": "参数不完整"}), 400
    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400
        br = _bedrock(ak, sk, region)
        if remove_tag_keys:
            br.untag_resource(resourceARN=profile_arn, tagKeys=remove_tag_keys)
        if tags_raw:
            br.tag_resource(
                resourceARN=profile_arn,
                tags=[{"key": k, "value": v} for k, v in tags_raw.items()],
            )
        return jsonify({"ok": True, "message": "标签已更新"})
    except Exception as e:
        return jsonify({"ok": False, "error": _aws_error(e)}), 400


@app.route("/api/delete_profile", methods=["POST"])
def delete_profile():
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    region = (data.get("region") or "").strip()
    profile_id = (data.get("inference_profile_arn") or data.get("inference_profile_id") or "").strip()
    if not all([ak, sk, region, profile_id]):
        return jsonify({"ok": False, "error": "参数不完整"}), 400
    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400
        br = _bedrock(ak, sk, region)
        br.delete_inference_profile(inferenceProfileIdentifier=profile_id)
        return jsonify({"ok": True, "message": "已删除"})
    except Exception as e:
        return jsonify({"ok": False, "error": _aws_error(e)}), 400


@app.route("/api/batch_list_profiles", methods=["POST"])
def batch_list_profiles():
    """多区域列出 APPLICATION inference profile（可按名称前缀过滤）"""
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    regions = [r.strip() for r in (data.get("regions") or []) if r and r.strip()]
    name_prefix = (data.get("name_prefix") or "").strip()
    if not all([ak, sk]):
        return jsonify({"ok": False, "error": "请填写 AK/SK"}), 400
    if not regions:
        return jsonify({"ok": False, "error": "请至少选择一个区域"}), 400
    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400
        profiles = []
        for region in regions:
            br = _bedrock(ak, sk, region)
            for p in _list_application_profiles(br, name_prefix or None):
                profiles.append({**p, "region": region})
        return jsonify({"ok": True, "total": len(profiles), "profiles": profiles})
    except Exception as e:
        return jsonify({"ok": False, "error": _aws_error(e)}), 400


@app.route("/api/batch_delete_profiles", methods=["POST"])
def batch_delete_profiles():
    """批量删除 inference profile（指定列表，或按区域 × 名称前缀扫描删除）"""
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    regions = [r.strip() for r in (data.get("regions") or []) if r and r.strip()]
    name_prefix = (data.get("name_prefix") or "").strip()
    items = data.get("profiles") or []

    if not all([ak, sk]):
        return jsonify({"ok": False, "error": "请填写 AK/SK"}), 400

    targets = []
    if items:
        for it in items:
            region = (it.get("region") or "").strip()
            pid = (it.get("inference_profile_arn") or it.get("inference_profile_id") or it.get("arn") or "").strip()
            if region and pid:
                targets.append({"region": region, "id": pid, "name": it.get("name", "")})
    elif name_prefix and regions:
        if len(name_prefix) < 2:
            return jsonify({"ok": False, "error": "名称前缀至少 2 个字符，避免误删"}), 400
        for region in regions:
            br = _bedrock(ak, sk, region)
            for p in _list_application_profiles(br, name_prefix):
                targets.append({
                    "region": region,
                    "id": p.get("arn") or p.get("id", ""),
                    "name": p.get("name", ""),
                })
    else:
        return jsonify({"ok": False, "error": "请指定要删除的配置列表，或填写名称前缀并选择区域"}), 400

    if not targets:
        return jsonify({"ok": True, "total": 0, "success": 0, "failed": 0, "results": [], "message": "未找到匹配的配置"})

    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400

        results = []
        clients = {}
        for t in targets:
            region = t["region"]
            pid = t["id"]
            if region not in clients:
                clients[region] = _bedrock(ak, sk, region)
            br = clients[region]
            try:
                br.delete_inference_profile(inferenceProfileIdentifier=pid)
                results.append({
                    "region": region,
                    "name": t.get("name", ""),
                    "inference_profile_arn": pid,
                    "ok": True,
                })
            except Exception as e:
                results.append({
                    "region": region,
                    "name": t.get("name", ""),
                    "inference_profile_arn": pid,
                    "ok": False,
                    "error": _aws_error(e),
                })

        ok_cnt = sum(1 for r in results if r["ok"])
        return jsonify({
            "ok": True,
            "total": len(results),
            "success": ok_cnt,
            "failed": len(results) - ok_cnt,
            "results": results,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": _aws_error(e)}), 500


@app.route("/api/tag_profiles", methods=["POST"])
def tag_profiles():
    """批量标签操作：支持添加/删除标签"""
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    region = (data.get("region") or "").strip()
    resource_arns = data.get("resource_arns", [])
    tags_raw = data.get("tags", {})
    remove_tag_keys = data.get("remove_tag_keys", [])
    if not all([ak, sk, region]):
        return jsonify({"ok": False, "error": "参数不完整"}), 400
    if not resource_arns:
        return jsonify({"ok": False, "error": "请选择配置"}), 400
    if not tags_raw and not remove_tag_keys:
        return jsonify({"ok": False, "error": "请添加或者选择要删除的标签"}), 400
    tag_list = [{"key": k, "value": v} for k, v in tags_raw.items()]
    results = []
    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400
        br = _bedrock(ak, sk, region)
        for arn in resource_arns:
            try:
                if remove_tag_keys:
                    br.untag_resource(resourceARN=arn, tagKeys=remove_tag_keys)
                if tag_list:
                    br.tag_resource(resourceARN=arn, tags=tag_list)
                results.append({"arn": arn, "ok": True})
            except Exception as e:
                results.append({"arn": arn, "ok": False, "error": str(e)})
        ok_cnt = sum(1 for r in results if r["ok"])
        fail_cnt = len(results) - ok_cnt
        return jsonify({
            "ok": True,
            "total": len(results),
            "success": ok_cnt,
            "failed": fail_cnt,
            "details": results,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/test_profile", methods=["POST"])
def test_profile():
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    region = (data.get("region") or "").strip()
    profile_arn = (data.get("inference_profile_arn") or "").strip()
    do_invoke = bool(data.get("invoke", True))
    check_source = bool(data.get("check_source", False))
    if not all([ak, sk, region, profile_arn]):
        return jsonify({"ok": False, "error": "参数不完整"}), 400
    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400
        br = _bedrock(ak, sk, region)
        resp = br.get_inference_profile(inferenceProfileIdentifier=profile_arn)
        p = resp.get("inferenceProfile", resp)
        status = p.get("status", "UNKNOWN")
        profile_arn = p.get("inferenceProfileArn", profile_arn)
        result = {
            "ok": True,
            "available": status == "ACTIVE",
            "status": status,
            "inferenceProfileId": p.get("inferenceProfileId", ""),
            "inferenceProfileArn": profile_arn,
            "name": p.get("inferenceProfileName", ""),
            "type": p.get("type", ""),
            "invoke_ok": None,
            "invoke_error": None,
        }
        
        # 检查源类型
        if check_source:
            try:
                models = p.get("models", [])
                source_info = None
                if models:
                    # 从 models 中获取 copyFrom 信息
                    first_model = models[0]
                    copy_from = first_model.get("modelArn", "")
                    if "inference-profile" in copy_from:
                        # 提取源 inference profile ID
                        source_id = copy_from.split("/")[-1] if "/" in copy_from else copy_from
                        source_type = _get_source_type(source_id)
                        source_info = {
                            "copyFrom": source_id,
                            "sourceType": source_type,
                            "isGlobal": source_type == "global"
                        }
                    else:
                        # Foundation model
                        source_info = {
                            "copyFrom": copy_from,
                            "sourceType": "foundation", 
                            "isGlobal": False
                        }
                
                if source_info:
                    result.update(source_info)
                else:
                    result.update({
                        "copyFrom": None,
                        "sourceType": "unknown",
                        "isGlobal": False
                    })
            except Exception as e:
                result.update({
                    "copyFrom": None,
                    "sourceType": "error",
                    "isGlobal": False,
                    "sourceError": str(e)
                })
        
        if do_invoke:
            if status != "ACTIVE":
                # Profile 本身不 ACTIVE，无需尝试调用
                result["invoke_ok"] = False
                result["invoke_error"] = f"Profile 状态为 {status}，无法调用"
                result["available"] = False
            else:
                try:
                    rt = _bedrock_runtime(ak, sk, region)
                    invoke_resp = rt.converse(
                        modelId=profile_arn,
                        messages=[{"role": "user", "content": [{"text": "hi"}]}],
                        inferenceConfig={"maxTokens": 8},
                    )
                    text = ""
                    for block in invoke_resp.get("output", {}).get("message", {}).get("content", []):
                        if "text" in block:
                            text += block["text"]
                    result["invoke_ok"] = True
                    result["invoke_preview"] = (text or "")[:80]
                    result["available"] = True   # 调用成功才算真正可用
                except Exception as e:
                    result["invoke_ok"] = False
                    result["invoke_error"] = _aws_error(e)
                    result["available"] = False  # 调用失败 = 不可用
        return jsonify(result)
    except botocore.exceptions.ClientError as e:
        code = e.response.get("Error", {}).get("Code", "")
        if code in ("ResourceNotFoundException", "NotFoundException"):
            return jsonify({"ok": True, "available": False, "status": "NOT_FOUND",
                            "error": "ARN 不存在或无访问权限"})
        return jsonify({"ok": False, "available": False, "error": _aws_error(e)})
    except Exception as e:
        return jsonify({"ok": False, "available": False, "error": _aws_error(e)})



@app.route("/api/export_excel", methods=["POST"])
def export_excel():
    """导出批量创建结果为格式化 Excel，带单元格合并。

    列顺序：
      A  AWS账户ID  ─ 全部数据行合并（同一账户）
      B  登录URL    ─ 全部数据行合并
      C  账号       ─ 全部数据行合并（留空，用户填）
      D  密码       ─ 全部数据行合并（留空，用户填）
      E  Access Key ─ 全部数据行合并
      F  Secret Key ─ 全部数据行合并
      G  模型类型   ─ 同一模型的连续行合并
      H  区域       ─ 每行独立
      I  模型ARN    ─ 每行独立
      J  标签       ─ 连续相同标签的行合并
    """
    data = request.get_json() or {}
    raw_results = data.get("results", [])
    account_id  = (data.get("account_id") or "").strip()
    access_key  = (data.get("access_key") or "").strip()
    secret_key  = (data.get("secret_key") or "").strip()

    if not raw_results:
        return jsonify({"ok": False, "error": "无数据可导出"}), 400

    # ── 预处理：计算每行的展示值 ──────────────────────────────
    login_url = (
        f"https://{account_id}.signin.aws.amazon.com/console"
        if account_id else ""
    )

    def _tags_str(r):
        if not r.get("ok"):
            return ""
        tags_raw = r.get("tags") or {}
        if isinstance(tags_raw, dict):
            return ", ".join(f"{k}={v}" for k, v in tags_raw.items())
        if isinstance(tags_raw, list):
            return ", ".join(
                f"{t.get('key','')}={t.get('value','')}" for t in tags_raw
            )
        return ""

    rows = []   # list of dicts, one per result row
    for r in raw_results:
        rows.append({
            "account_id":  account_id,
            "login_url":   login_url,
            "username":    "",           # C 留空
            "password":    "",           # D 留空
            "access_key":  access_key,   # E
            "secret_key":  secret_key,   # F
            "model_label": r.get("model_label", r.get("model_id", "")),
            "region":      r.get("region", ""),
            "arn":         r.get("inferenceProfileArn", "") if r.get("ok")
                           else r.get("error", ""),
            "tags":        _tags_str(r),
            "ok":          bool(r.get("ok")),
        })

    # ── 按模型类型排序，相同模型的行聚在一起，再按区域排序 ──
    rows.sort(key=lambda x: (x["model_label"], x["region"]))

    # ── 工作簿 & 样式 ────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Inference Profiles"

    # 表头：不加粗，居中
    hdr_font  = Font(name="Calibri", size=11, bold=False)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 普通数据：左对齐
    val_align   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    # 居中对齐（合并列、账号密码、区域/模型/标签）
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _border(top="thin", bottom="thin", left="thin", right="thin"):
        def _side(s):
            return Side(style=s) if s else Side(style=None)
        return Border(left=_side(left), right=_side(right),
                      top=_side(top),   bottom=_side(bottom))

    full_border = _border()
    err_font    = Font(name="Calibri", italic=True)

    # ── 表头（行 1） ─────────────────────────────────────────
    headers = ["AWS账户ID", "登录URL", "账号", "密码", "Access Key", "Secret Key",
               "模型类型", "区域", "模型ARN", "标签"]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = full_border
    ws.row_dimensions[1].height = 20

    # ── 写入数据（行 2 起） ──────────────────────────────────
    data_start = 2
    total_rows = len(rows)

    # 居中列：C(3)账号  D(4)密码  G(7)模型类型  H(8)区域  J(10)标签
    CENTER_COLS = {3, 4, 7, 8, 10}

    for ri, row in enumerate(rows):
        excel_row = data_start + ri
        values = [
            row["account_id"],  # A 1
            row["login_url"],   # B 2
            row["username"],    # C 3
            row["password"],    # D 4
            row["access_key"],  # E 5
            row["secret_key"],  # F 6
            row["model_label"], # G 7
            row["region"],      # H 8
            row["arn"],         # I 9
            row["tags"],        # J 10
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=ci, value=val)
            cell.border    = full_border
            cell.alignment = center_align if ci in CENTER_COLS else val_align
            if ci == 9 and not row["ok"]:
                cell.font = err_font

    # ── 列宽：按内容自动计算，账号(C)/密码(D) 至少预留 20 ──────
    # 固定最小宽度（保证空列可用）
    MIN_WIDTHS = {3: 20, 4: 20}   # C=账号, D=密码

    col_letters = [chr(64 + i) for i in range(1, len(headers) + 1)]
    for ci, col_letter in enumerate(col_letters, 1):
        max_len = len(headers[ci - 1])
        for row in rows:
            cell_values = [
                row["account_id"], row["login_url"], row["username"],
                row["password"],   row["access_key"], row["secret_key"],
                row["model_label"], row["region"], row["arn"], row["tags"],
            ]
            cell_val = str(cell_values[ci - 1]) if cell_values[ci - 1] else ""
            max_len = max(max_len, len(cell_val))
        width = min(max_len + 4, 80)
        width = max(width, MIN_WIDTHS.get(ci, 0))
        ws.column_dimensions[col_letter].width = width

    # ── 合并辅助函数 ─────────────────────────────────────────
    def _apply_merge(ws, r1, r2, col, value, align, fill=None, font=None):
        """合并 (r1, col) ~ (r2, col)，写入值并补边框。"""
        if r1 == r2:
            # 单行不需要合并，但确保对齐已经在写入时设置
            return
        ws.merge_cells(
            start_row=r1, start_column=col,
            end_row=r2,   end_column=col
        )
        top_cell = ws.cell(row=r1, column=col, value=value)
        top_cell.alignment = align
        if fill:
            top_cell.fill = fill
        if font:
            top_cell.font = font
        # 合并区域外围补全边框（openpyxl 合并后内部格会丢边框）
        for r in range(r1, r2 + 1):
            top    = "thin" if r == r1 else None
            bottom = "thin" if r == r2 else None
            ws.cell(row=r, column=col).border = _border(
                top=top, bottom=bottom, left="thin", right="thin"
            )

    # ── A/B/C/D/E/F：全部数据行合并 ─────────────────────────
    last_data = data_start + total_rows - 1
    if total_rows > 1:
        _apply_merge(ws, data_start, last_data, 1, account_id, center_align)
        _apply_merge(ws, data_start, last_data, 2, login_url,  center_align)
        _apply_merge(ws, data_start, last_data, 3, "",         center_align)
        _apply_merge(ws, data_start, last_data, 4, "",         center_align)
        _apply_merge(ws, data_start, last_data, 5, access_key, center_align)
        _apply_merge(ws, data_start, last_data, 6, secret_key, center_align)

    # ── G（模型类型）：同一模型连续行合并 ────────────────────
    i = 0
    while i < total_rows:
        j = i + 1
        while j < total_rows and rows[j]["model_label"] == rows[i]["model_label"]:
            j += 1
        r1, r2 = data_start + i, data_start + j - 1
        _apply_merge(ws, r1, r2, 7, rows[i]["model_label"], center_align)
        i = j

    # ── J（标签）：连续相同标签行合并 ────────────────────────
    i = 0
    while i < total_rows:
        j = i + 1
        while j < total_rows and rows[j]["tags"] == rows[i]["tags"]:
            j += 1
        r1, r2 = data_start + i, data_start + j - 1
        _apply_merge(ws, r1, r2, 10, rows[i]["tags"], center_align)
        i = j

    # ── 冻结首行 ─────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── 输出 ─────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    date_str = datetime.now().strftime("%Y%m%d")
    filename = (
        f"AWS-Bedrock-Arn-{account_id}-{date_str}.xlsx"
        if account_id
        else f"AWS-Bedrock-Arn-{date_str}.xlsx"
    )
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

@app.route("/quotas")
def quotas_page():
    return render_template("quotas.html")


@app.route("/api/export_quotas_excel", methods=["POST"])
def export_quotas_excel():
    """将配额查询结果导出为格式化 Excel"""
    data       = request.get_json(force=True) or {}
    rows       = data.get("rows") or []
    account_id = (data.get("account_id") or "").strip()
    use_global = bool(data.get("use_global", False))

    if not rows:
        return jsonify({"ok": False, "error": "无数据"}), 400

    # ── 解析并聚合：(region, model) → {tpd, tpm} ──────────────
    def _parse_name(name):
        n = name or ""
        m = re.search(r"\bfor\s+(.+)$", n, re.I)
        model = m.group(1).strip() if m else n
        model = re.sub(r"(?i)^anthropic\s+", "", model).strip()
        nl = n.lower()
        if   re.search(r"tokens per day",      nl): qtype = "TPD"
        elif re.search(r"tokens per minute",   nl): qtype = "TPM"
        elif re.search(r"requests per minute", nl): qtype = "RPM"
        else: qtype = "OTHER"
        return model, qtype

    agg = {}
    region_order = []
    for row in rows:
        region = row.get("region", "")
        model, qtype = _parse_name(row.get("name", ""))
        key = (region, model)
        if key not in agg:
            agg[key] = {"tpd": None, "tpm": None}
            if region not in region_order:
                region_order.append(region)
        if qtype == "TPD":
            agg[key]["tpd"] = row.get("value")
        elif qtype == "TPM":
            agg[key]["tpm"] = row.get("value")

    sorted_keys = []
    for region in region_order:
        region_keys = sorted([k for k in agg if k[0] == region], key=lambda x: x[1])
        sorted_keys.extend(region_keys)

    # ── 构建 Excel ────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "配额汇总"

    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin   = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    num_fmt = '#,##0'

    # ── 第1行：列标题 ─────────────────────────────────────────
    headers = ["AWS ID", "区域", "模型名称", "TPD", "TPM"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True)
        c.alignment = center
        c.border    = border
    ws.row_dimensions[1].height = 18

    # ── 数据行 ────────────────────────────────────────────────
    for i, key in enumerate(sorted_keys):
        region, model = key
        vals = agg[key]
        r = i + 2  # 数据从第2行开始

        # AWS ID 每行都写值，最后统一合并
        ws.cell(row=r, column=1, value=account_id).border = border
        ws.cell(row=r, column=1).alignment = center

        # 区域只在该区域第一条显示
        region_first = (i == 0 or sorted_keys[i-1][0] != region)
        ws.cell(row=r, column=2, value=region if region_first else "").border = border
        ws.cell(row=r, column=2).alignment = left

        ws.cell(row=r, column=3, value=model).border = border
        ws.cell(row=r, column=3).alignment = left

        for col, val in [(4, vals["tpd"]), (5, vals["tpm"])]:
            c = ws.cell(row=r, column=col)
            c.value = val if val is not None else "—"
            c.alignment = center
            c.border = border
            if val is not None:
                c.number_format = num_fmt

        ws.row_dimensions[r].height = 16

    # ── 合并 AWS ID 列（全部数据行合并为一格）────────────────
    total_data = len(sorted_keys)
    if total_data > 1:
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=1 + total_data, end_column=1)
        merged = ws["A2"]
        merged.alignment = center
        merged.border    = border

    # ── 合并区域列（相同区域的行合并）───────────────────────
    if sorted_keys:
        merge_start = 2
        cur_region  = sorted_keys[0][0]
        for i in range(1, len(sorted_keys)):
            if sorted_keys[i][0] != cur_region:
                end_row = i + 1        # excel row = i(0-based)+2-1 = i+1
                if end_row >= merge_start:
                    ws.merge_cells(start_row=merge_start, start_column=2,
                                   end_row=end_row,        end_column=2)
                    c = ws.cell(row=merge_start, column=2)
                    c.alignment = left
                    c.border    = border
                merge_start = i + 2
                cur_region  = sorted_keys[i][0]
        # 最后一组
        last_row = len(sorted_keys) + 1
        if last_row >= merge_start:
            ws.merge_cells(start_row=merge_start, start_column=2,
                           end_row=last_row,      end_column=2)
            c = ws.cell(row=merge_start, column=2)
            c.alignment = left
            c.border    = border

    # ── 列宽 ─────────────────────────────────────────────────
    for i, w in enumerate([16, 16, 38, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import datetime as _dt
    ts = _dt.now().strftime("%Y%m%d")
    fname = f"bedrock-quotas-{account_id}-{ts}.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── 持久化 code_map（写到 outputs/quota_codes.json，跨重启复用）────
import json as _json
import os as _os
import time as _time
import concurrent.futures as _cf

_CODE_MAP_PATH = _os.path.join(_os.path.dirname(__file__), "outputs", "quota_codes.json")
_quota_code_map: dict = {}   # lower(name) → QuotaCode，全局内存
_quota_code_loaded = False

def _load_code_map_from_disk():
    global _quota_code_map, _quota_code_loaded
    if _quota_code_loaded:
        return
    try:
        if _os.path.exists(_CODE_MAP_PATH):
            with open(_CODE_MAP_PATH, "r", encoding="utf-8") as f:
                _quota_code_map = _json.load(f)
    except Exception:
        pass
    _quota_code_loaded = True

def _save_code_map_to_disk():
    try:
        _os.makedirs(_os.path.dirname(_CODE_MAP_PATH), exist_ok=True)
        with open(_CODE_MAP_PATH, "w", encoding="utf-8") as f:
            _json.dump(_quota_code_map, f, ensure_ascii=False)
    except Exception:
        pass

def _build_code_map(ak, sk):
    """用 us-east-1 list_aws_default_service_quotas 建立 code_map，写入磁盘"""
    global _quota_code_map
    new_map = {}
    try:
        client = _get_client("service-quotas", ak, sk, "us-east-1")
        for page in client.get_paginator("list_aws_default_service_quotas").paginate(ServiceCode="bedrock"):
            for q in page.get("Quotas", []):
                name = q.get("QuotaName", "")
                nl   = name.lower()
                if ("claude" in nl or "anthropic" in nl) and q.get("QuotaCode"):
                    new_map[nl] = q["QuotaCode"]
    except Exception:
        pass
    if new_map:
        _quota_code_map = new_map
        _save_code_map_to_disk()


def _refresh_code_map(ak, sk):
    """强制重新从 AWS 拉取最新配额代码（忽略缓存）"""
    global _quota_code_map, _quota_code_loaded
    _quota_code_loaded = False
    _quota_code_map = {}
    _build_code_map(ak, sk)

# 应用启动时从磁盘加载
_load_code_map_from_disk()


@app.route("/api/query_quotas", methods=["POST"])
def query_quotas():
    """查询 Bedrock Claude 配额 — 支持多区域并发"""
    import concurrent.futures as _cf2
    data        = request.get_json(force=True) or {}
    ak, sk, _   = _creds(data)
    regions_raw = data.get("regions") or ([data.get("region")] if data.get("region") else [])
    regions     = [r.strip() for r in regions_raw if r and r.strip()]
    quota_types = data.get("quota_types") or []
    sel_models  = data.get("models") or []
    use_global  = bool(data.get("use_global", False))
    force_refresh = bool(data.get("force_refresh", False))

    if not ak or not sk:
        return jsonify({"ok": False, "error": "请提供 access_key 和 secret_key"})
    if not regions:
        return jsonify({"ok": False, "error": "请选择区域"})

    TYPE_KEYWORDS = {
        "tpm": "tokens per minute",
        "tpd": "tokens per day",
        "rpm": "requests per minute",
    }
    type_filters = [TYPE_KEYWORDS[t.lower()] for t in quota_types if t.lower() in TYPE_KEYWORDS]
    EXCLUDE_KEYWORDS = ["on-demand", "doubled for cross-region", "[bedrock-mantle endpoint]",
                        "latency-optimized", "provisioned", "batch inference", "customization",
                        "minimum number", "records per", "sum of in-progress"]

    # 强制刷新或按需加载 code_map
    if force_refresh:
        _refresh_code_map(ak, sk)
    else:
        _load_code_map_from_disk()
        if not _quota_code_map:
            _build_code_map(ak, sk)
            if not _quota_code_map:
                return jsonify({"ok": False, "error": "无法获取配额代码映射，请稍后重试"})

    # 把模型 label 转成更宽泛的关键词进行匹配
    # 例如 "Claude Sonnet 5" → ["sonnet 5"]
    # 例如 "Claude Sonnet 4.5" → ["sonnet 4.5"]
    def _model_keywords(label):
        """从 label 提取 quota name 里会出现的关键词"""
        s = label.lower().replace("claude ", "").strip()  # "sonnet 5", "haiku 4.5", ...
        return [s]

    model_kws = []
    for m in sel_models:
        model_kws.extend(_model_keywords(m))

    # 筛选目标配额代码
    targets = []
    for nl, code in _quota_code_map.items():
        # 基础过滤：排除不需要的配额类型
        if any(ex in nl for ex in EXCLUDE_KEYWORDS):
            continue
        # 模型过滤：任一关键词匹配即可
        if model_kws and not any(kw in nl for kw in model_kws):
            continue
        # 配额类型过滤
        if type_filters and not any(kw in nl for kw in type_filters):
            continue
        # global/cross-region 去重
        if use_global:
            if nl.startswith("cross-region "):
                if ("global " + nl) in _quota_code_map:
                    continue
        else:
            if nl.startswith("global cross-region "):
                if nl[len("global "):] in _quota_code_map:
                    continue
        targets.append((nl, code))

    if not targets:
        # 返回当前 code_map 里所有 claude 相关的配额名，方便调试
        sample = [k for k in list(_quota_code_map.keys())[:10]]
        return jsonify({
            "ok": True, "quotas": [], "errors": [],
            "total": 0, "use_global": use_global,
            "debug_model_kws": model_kws,
            "debug_map_sample": sample,
            "debug_map_size": len(_quota_code_map),
        })

    sc = "bedrock"
    all_results = []
    all_errors  = []

    def _query_region(region):
        client = _get_client("service-quotas", ak, sk, region)
        region_results = []

        # 先用第一个 code 试探权限
        if targets:
            test_nl, test_code = targets[0]
            try:
                client.get_aws_default_service_quota(ServiceCode=sc, QuotaCode=test_code)
            except botocore.exceptions.ClientError as e:
                code = e.response.get("Error", {}).get("Code", "")
                if code in ("AccessDeniedException", "AccessDenied", "AuthorizationError"):
                    raise PermissionError(f"账号无权限查询 Service Quotas（{region}），请确认 IAM 权限包含 servicequotas:GetServiceQuota")
            except Exception:
                pass  # 其他错误继续尝试

        def _fetch_one(nl_code):
            nl, code = nl_code
            aq, dq = {}, {}
            try:
                aq = client.get_service_quota(ServiceCode=sc, QuotaCode=code).get("Quota", {})
            except botocore.exceptions.ClientError as e:
                err_code = e.response.get("Error", {}).get("Code", "")
                if err_code in ("AccessDeniedException", "AccessDenied", "AuthorizationError"):
                    raise PermissionError(f"无权限: {e.response['Error']['Message']}")
            except Exception:
                pass
            try:
                dq = client.get_aws_default_service_quota(ServiceCode=sc, QuotaCode=code).get("Quota", {})
            except Exception:
                pass
            if not aq and not dq:
                return None
            return {
                "region":        region,
                "name":          (aq or dq).get("QuotaName", ""),
                "value":         aq.get("Value"),
                "default_value": dq.get("Value"),
                "quota_code":    code,
            }

        max_w = min(30, len(targets))
        with _cf.ThreadPoolExecutor(max_workers=max_w) as pool:
            futures = {pool.submit(_fetch_one, t): t for t in targets}
            for future in _cf.as_completed(futures):
                try:
                    item = future.result()
                    if item:
                        region_results.append(item)
                except PermissionError:
                    raise   # 权限错误向上传递
                except Exception:
                    pass
        region_results.sort(key=lambda x: x["name"])
        return region_results

    max_region_workers = min(10, len(regions))
    with _cf.ThreadPoolExecutor(max_workers=max_region_workers) as pool:
        future_map = {pool.submit(_query_region, r): r for r in regions}
        for future in _cf.as_completed(future_map):
            region = future_map[future]
            try:
                rows = future.result()
                all_results.extend(rows)
            except PermissionError as e:
                all_errors.append({
                    "region": region,
                    "error": str(e),
                    "type": "permission"   # 前端用于区分展示
                })
            except Exception as e:
                all_errors.append({"region": region, "error": str(e), "type": "error"})

    all_results.sort(key=lambda x: (x["region"], x["name"]))
    return jsonify({
        "ok":         True,
        "quotas":     all_results,
        "errors":     all_errors,
        "total":      len(all_results),
        "use_global": use_global,
        "regions":    regions,
    })


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/model-tags")
def model_tags():
    return render_template("model_tags.html")


@app.route("/delete-profiles")
def delete_profiles():
    return render_template("delete_profiles.html")


@app.route("/data-retention")
def data_retention_page():
    return render_template("data_retention.html")


@app.route("/api/get_data_retention", methods=["POST"])
def get_data_retention():
    """获取当前数据驻留设置"""
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    region = (data.get("region") or "us-east-1").strip()
    
    if not all([ak, sk, region]):
        return jsonify({"ok": False, "error": "请填写必要参数"}), 400
        
    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400
                
        # 使用 requests 库调用 Bedrock 数据驻留 API
        import requests
        from botocore.auth import SigV4Auth
        from botocore.awsrequest import AWSRequest
        
        # 构建请求
        endpoint = f"https://bedrock.{region}.amazonaws.com/data-retention"
        request_obj = AWSRequest(method="GET", url=endpoint)
        
        # 签名请求
        credentials = boto3.Session(
            aws_access_key_id=ak,
            aws_secret_access_key=sk
        ).get_credentials()
        
        SigV4Auth(credentials, "bedrock", region).add_auth(request_obj)
        
        # 发送请求
        response = requests.get(
            endpoint,
            headers=dict(request_obj.headers)
        )
        
        if response.status_code == 200:
            result = response.json()
            return jsonify({
                "ok": True,
                "mode": result.get("mode", "unknown"),
                "updatedAt": result.get("updatedAt", "")
            })
        else:
            return jsonify({
                "ok": False, 
                "error": f"API 调用失败: {response.status_code} - {response.text}"
            }), 400
            
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/api/set_data_retention", methods=["POST"])
def set_data_retention():
    """设置数据驻留模式"""
    data = request.get_json() or {}
    ak, sk, user_id = _creds(data)
    region = (data.get("region") or "us-east-1").strip()
    mode = (data.get("mode") or "").strip()
    
    if not all([ak, sk, region, mode]):
        return jsonify({"ok": False, "error": "请填写必要参数"}), 400
        
    if mode not in ["provider_data_share", "no_data_share"]:
        return jsonify({"ok": False, "error": "无效的数据驻留模式"}), 400
        
    try:
        if user_id:
            _, err = _verify_account(ak, sk, user_id)
            if err:
                return jsonify({"ok": False, "error": err}), 400
                
        # 使用 requests 库调用 Bedrock 数据驻留 API
        import requests
        from botocore.auth import SigV4Auth
        from botocore.awsrequest import AWSRequest
        import json as json_lib
        
        # 构建请求
        endpoint = f"https://bedrock.{region}.amazonaws.com/data-retention"
        payload = json_lib.dumps({"mode": mode})
        
        request_obj = AWSRequest(
            method="PUT", 
            url=endpoint,
            data=payload,
            headers={"Content-Type": "application/json"}
        )
        
        # 签名请求
        credentials = boto3.Session(
            aws_access_key_id=ak,
            aws_secret_access_key=sk
        ).get_credentials()
        
        SigV4Auth(credentials, "bedrock", region).add_auth(request_obj)
        
        # 发送请求
        response = requests.put(
            endpoint,
            data=payload,
            headers=dict(request_obj.headers)
        )
        
        if response.status_code == 200:
            result = response.json()
            return jsonify({
                "ok": True,
                "mode": result.get("mode"),
                "updatedAt": result.get("updatedAt", "")
            })
        else:
            return jsonify({
                "ok": False, 
                "error": f"API 调用失败: {response.status_code} - {response.text}"
            }), 400
            
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/test-profile")
def test_profile_page():
    return render_template("test_profile.html")


@app.route("/mfa")
def mfa_page():
    return render_template("mfa.html")


# ── MFA / TOTP ──────────────────────────────────────────────────────────────

@app.route("/api/mfa/totp", methods=["POST"])
def mfa_totp():
    """根据 Base32 密钥计算当前 TOTP 验证码及剩余秒数"""
    import pyotp, time as _t
    data   = request.get_json(force=True) or {}
    secret = (data.get("secret") or "").strip().upper().replace(" ", "")
    if not secret:
        return jsonify({"ok": False, "error": "请提供密钥"})
    try:
        totp   = pyotp.TOTP(secret)
        code   = totp.now()
        remain = 30 - int(_t.time()) % 30
        return jsonify({"ok": True, "code": code, "remain": remain})
    except Exception as e:
        return jsonify({"ok": False, "error": f"密钥格式错误：{e}"})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=False, use_reloader=False, threaded=True)
